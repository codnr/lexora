#Name:          Lexora EDI
#Description:   Automate EDI
#Author:        Codnr
#Version:       1
#Require:       openpyxl==2.5.14
#Notice:        Close all files used by this program before runing.

import openpyxl

def GetSheetbyName(Spreadsheet,Name):
    return Spreadsheet.sheetnames[Spreadsheet.sheetnames.index(Name)]

def main():
    #Read
    LexoraSpreadSheet = openpyxl.load_workbook("LEXORA Master DATA FILE_FOR_DEALERS_6-25-19.xlsx")
    #print(LexoraSpreadSheet.sheetnames)
    #BathroomVanities = LexoraSpreadSheet['Bathroom Vanities']
    #BathtubsandFaucets = LexoraSpreadSheet['Bathtubs&Faucets']
    #Fireplaces = LexoraSpreadSheet['Fireplaces']
    #UpdateHistory = LexoraSpreadSheet['Update History']
    Fireplaces = LexoraSpreadSheet['Fireplaces']
    #print(Fireplaces)
    #print(type(Fireplaces))
    RUPC = str(Fireplaces['E2'].value)
    print('RUPC: ',RUPC)
    #Read


    #Write
    LexoraLowesSpreadSheet = openpyxl.load_workbook("LexoraHome-Lowe's-Lowe's-CoreMarketing-Lowe's-BathroomVanitieswith-08082019.xlsx")
    #Summary = LexoraLowesSpreadSheet['Summary']
    Attributes = LexoraLowesSpreadSheet['Attributes']
    #Reference = LexoraLowesSpreadSheet['Reference']
    WUPC = Attributes['DW6'].value
    if WUPC == None:
        Attributes['DW6'] = RUPC
        print('Written')
        print('WUPC: ',Attributes['DW6'].value)
    else:
        print('Existing Value')
        print('WUPC: ',WUPC)
    LexoraLowesSpreadSheet.save(filename = "LexoraHome-Lowe's-Lowe's-CoreMarketing-Lowe's-BathroomVanitieswith-08082019.xlsx")
    #Write


if __name__ == '__main__':
    main()