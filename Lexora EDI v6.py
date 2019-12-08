#Name:          Lexora EDI
#Description:   Automate EDI
#Author:        Codnr
#Version:       6
#Require:       openpyxl==2.5.14
#Notice:        Close all files used by this program before runing.

import openpyxl

def GenerateMapping():
    Mapping = []
    Data = []
    Data.append("UPC")#Name
    Data.append(["UPC"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("GTIN")#Name
    Data.append(["GTIN","globalTradeItemNumber (GTIN)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Highest Level GTIN")#Name
    Data.append(["Highest Level GTIN"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Product Name")#Name
    Data.append(["Title","Product Name"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Model Number")#Name
    Data.append(["Model Number","MFG Part # (OEM)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vanity Color")#Name
    Data.append(["Vanity Color","Manufacturer Color/Finish"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Color/Finish Family")#Name
    Data.append(["Color/Finish Family"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Mirror Included")#Name
    Data.append(["Mirror","Mirror Included"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Collection")#Name
    Data.append(["Collection","Collection Name"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Product Type")#Name
    Data.append(["Product Type"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Counter")#Name
    Data.append(["Counter"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Counter Color")#Name
    Data.append(["Counter Color","Manufacturer Top Color"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Information Provider GLN")#Name
    Data.append(["Information Provider GLN"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("0810014570006")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Base Warranty Parts")#Name
    Data.append(["Base Warranty Parts"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("1-year limited")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Base Warranty Labor")#Name
    Data.append(["Base Warranty Labor"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("1-year limited")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Dropship Item?")#Name
    Data.append(["Dropship Item?"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("California Proposition 65 Warning Required")#Name
    Data.append(["California Proposition 65 Warning Required"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("California Prop 65 Warning Label (Digital Asset)")#Name
    Data.append(["California Prop 65 Warning Label (Digital Asset)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("CaliforniaProposition65Compliance.pdf")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Item has Restrictions?")#Name
    Data.append(["Item has Restrictions?"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Lowe's Vendor Type")#Name
    Data.append(["Lowe's Vendor Type"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("LGS/IOS Import Vendor")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Complies with Federal Lead Guidelines")#Name
    Data.append(["Complies with Federal Lead Guidelines"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("CARB Compliant")#Name
    Data.append(["CARB Compliant"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Solid Wood Frame")#Name
    Data.append(["Solid Wood Frame"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Soft Close Doors")#Name
    Data.append(["Soft Close Doors"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Integrated Electrical Outlet")#Name
    Data.append(["Integrated Electrical Outlet"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Overflow Drain Included")#Name
    Data.append(["Overflow Drain Included"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Hardware Finish")#Name
    Data.append(["Hardware Finish"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Chrome")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Manufacturer Faucet Finish")#Name
    Data.append(["Manufacturer Faucet Finish"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("N/A")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Soft Close Drawers")#Name
    Data.append(["Soft Close Drawers"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Medicine Cabinet Included")#Name
    Data.append(["Medicine Cabinet Included"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Side Panel Material")#Name
    Data.append(["Side Panel Material"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("N/A")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Top Material")#Name
    Data.append(["Top Material"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Marble")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Item has Restrictions?")#Name
    Data.append(["Item has Restrictions?"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Style")#Name
    Data.append(["Style"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Modern")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Sink Shape")#Name
    Data.append(["Sink Shape"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Square")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Modular Design")#Name
    Data.append(["Modular Design"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Sink Material")#Name
    Data.append(["Sink Material"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Ceramic")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Edge Profile")#Name
    Data.append(["Edge Profile"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Flat")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Faucet Mount Type")#Name
    Data.append(["Faucet Mount Type"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Single hole")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Unfinished Sides")#Name
    Data.append(["Unfinished Sides"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("None")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Adjustable Shelves")#Name
    Data.append(["Adjustable Shelves"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Warranty")#Name
    Data.append(["Warranty"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("1-year limited")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Decorative Hardware Included")#Name
    Data.append(["Decorative Hardware Included"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Max Flow Rate")#Name
    Data.append(["Max Flow Rate"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("N/A")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Sink Offset")#Name
    Data.append(["Sink Offset"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("N/A")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Hardware Color Family")#Name
    Data.append(["Hardware Color Family"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Chrome")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Dove Tail Drawer Construction")#Name
    Data.append(["Dove Tail Drawer Construction"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("N/A")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Faucet Included")#Name
    Data.append(["Faucet Included"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Interior Color")#Name
    Data.append(["Interior Color"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Matches exterior")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vanity Top Thickness in Inches")#Name
    Data.append(["Vanity Top Thickness in Inches"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("3/4-in")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Toe Kick")#Name
    Data.append(["Toe Kick"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Lowe's Exclusive")#Name
    Data.append(["Lowe's Exclusive"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Import Home Office VBU - CAN")#Name
    Data.append(["Import Home Office VBU - CAN"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("103831")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Import Home Office VBU - USA")#Name
    Data.append(["Import Home Office VBU - USA"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("103831")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Factory Name")#Name
    Data.append(["Factory Name"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Lexora")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Factory Party ID")#Name
    Data.append(["Factory Party ID"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("0")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("GTIN Wood Percentage")#Name
    Data.append(["GTIN Wood Percentage"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("100%")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("TSCA Title VI")#Name
    Data.append(["TSCA Title VI"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Product contains no composite wood")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Hazardous Indicator")#Name
    Data.append(["Hazardous Indicator"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Component Material (IOS)")#Name
    Data.append(["Component Material (IOS)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("80% rubber wood and 20% birch")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Does this item include any textile materials?")#Name
    Data.append(["Does this item include any textile materials?"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("In what country was the fabric cut?")#Name
    Data.append(["In what country was the fabric cut?"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("CHINA")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("In what country was the fabric formed? (i.e., knit, woven, etc.)")#Name
    Data.append(["In what country was the fabric formed? (i.e., knit, woven, etc.)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("CHINA")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("In what country was the fabric sewn?")#Name
    Data.append(["In what country was the fabric sewn?"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("CHINA")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Is the fabric knit, woven, or non-woven?")#Name
    Data.append(["Is the fabric knit, woven, or non-woven?"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Non-Woven")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("What is the name for the mill where the fabric was formed? (i.e., knit, woven, etc.)")#Name
    Data.append(["What is the name for the mill where the fabric was formed? (i.e., knit, woven, etc.)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("N/A")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Mill Address")#Name
    Data.append(["Mill Address"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("N/A")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Does this product consist of multiple items (set)?")#Name
    Data.append(["Does this product consist of multiple items (set)?"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Origin Country")#Name
    Data.append(["Origin Country"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("CHINA")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Shipping Port (Code)")#Name
    Data.append(["Shipping Port (Code)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("USEWR")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Shipping Port Description")#Name
    Data.append(["Shipping Port Description"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("NEWARK")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Custom HTS Code (IOS)")#Name
    Data.append(["Custom HTS Code (IOS)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("9403604000")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Item Description (IOS)")#Name
    Data.append(["Item Description (IOS)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("N/A")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Import Builder Shipping Type")#Name
    Data.append(["Import Builder Shipping Type"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("NOT APPLICABLE")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Supporting Sourcing Office Location")#Name
    Data.append(["Supporting Sourcing Office Location"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Shanghai")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Ship from VBU (Ship from Factory)")#Name
    Data.append(["Ship from VBU (Ship from Factory)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("103831")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vendor Contact Name")#Name
    Data.append(["Vendor Contact Name"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Lexora")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vendor Email Address")#Name
    Data.append(["Vendor Email Address"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("general@lexorahome.com")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vendor Name")#Name
    Data.append(["Vendor Name"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Lexora")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vendor Phone Number")#Name
    Data.append(["Vendor Phone Number"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("855-453-9672")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("MFG Name")#Name
    Data.append(["MFG Name"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Lexora Home")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Manufacturer Address")#Name
    Data.append(["Manufacturer Address"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("425 Ferry Street, Newark, NJ 07105")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Lowes Merchandising Contact Name")#Name
    Data.append(["Lowes Merchandising Contact Name"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Amanda L. Wall")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Lowes Merchandising Contact Email")#Name
    Data.append(["Lowes Merchandising Contact Email"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Amanda.L.Wall@lowes.com")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Item Description {CAN_DOM}")#Name
    Data.append(["Item Description {CAN_DOM}"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("N/A")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Custom HTS Code (CA Dom)")#Name
    Data.append(["Custom HTS Code (CA Dom)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("9403604000")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Component Material")#Name
    Data.append(["Component Material"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("80% rubber wood and 20% birch.")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Country Of Origin")#Name
    Data.append(["Country Of Origin"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("CHINA")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Is Product NAFTA Eligible (Made in USA or Mexico)")#Name
    Data.append(["Is Product NAFTA Eligible (Made in USA or Mexico)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("What is the product’s function")#Name
    Data.append(["What is the product’s function"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Bathroom Vanity")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 1")#Name
    Data.append(["Bullet point 1","Feature - Benefit Bullet 1"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("8-stage painting and finishing procs, each finish is primed and sealed for superior moisture resistance")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 2")#Name
    Data.append(["Bullet point 2","Feature - Benefit Bullet 2"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Solid rubber and birch wood")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 3")#Name
    Data.append(["Bullet point 3","Feature - Benefit Bullet 3"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 4")#Name
    Data.append(["Bullet point 4","Feature - Benefit Bullet 4"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 5")#Name
    Data.append(["Bullet point 5","Feature - Benefit Bullet 5"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 6")#Name
    Data.append(["Bullet point 6","Feature - Benefit Bullet 6"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 7")#Name
    Data.append(["Bullet point 7","Feature - Benefit Bullet 7"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 8")#Name
    Data.append(["Bullet point 8","Feature - Benefit Bullet 8"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 9")#Name
    Data.append(["Bullet point 9","Feature - Benefit Bullet 9"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 10")#Name
    Data.append(["Bullet point 10"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 11")#Name
    Data.append(["Bullet point 11"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 12")#Name
    Data.append(["Bullet point 12"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 13")#Name
    Data.append(["Bullet point 13"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 14")#Name
    Data.append(["Bullet point 14"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 15")#Name
    Data.append(["Bullet point 15"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Bullet Point 16")#Name
    Data.append(["Bullet point 16"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Net Weight Unpackaged")#Name
    Data.append(["ITEM NET Weight (w/o packaging of any kind) in Pounds"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Item Net Weight")#Name
    Data.append(["Item Net Weight"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vanity Weight Only")#Name
    Data.append(["Vanity Weight Only"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Counter Weight Only")#Name
    Data.append(["Counter Weight Only"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Mirror Weight Only")#Name
    Data.append(["Mirror Weight Only"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Weight")#Name
    Data.append(["Weight"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Product Dims")#Name
    Data.append(["Product Dims"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Cabinet Dims")#Name
    Data.append(["Vanity/Side Cabinet Dims"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Sink Dims")#Name
    Data.append(["Sink Dims"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Common Height")#Name
    Data.append(["Common Height"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Common Width")#Name
    Data.append(["Common Width"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Common Depth")#Name
    Data.append(["Common Depth"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Cabinet Height")#Name
    Data.append(["Cabinet Height"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Cabinet Width")#Name
    Data.append(["Cabinet Width"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Cabinet Depth")#Name
    Data.append(["Cabinet Depth"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Actual Depth")#Name
    Data.append(["Actual Depth","Actual Depth with Top"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Actual Height")#Name
    Data.append(["Actual Height","Actual Height with Top"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Actual Width")#Name
    Data.append(["Actual Width","Actual Width with Top"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vanity Width")#Name
    Data.append(["Vanity Width"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Compatible Vanity Top Width")#Name
    Data.append(["Compatible Vanity Top Width"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Compatible Vanity Top Depth")#Name
    Data.append(["Compatible Vanity Top Depth"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 1")#Name
    Data.append(["Image White Backgroud","Main Product Image"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 2")#Name
    Data.append(["Image White Backgroud no Logo","Detailed Product View 2"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 3")#Name
    Data.append(["Image 1","Detailed Product View 3"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 4")#Name
    Data.append(["Image 2","Detailed Product View 4"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 5")#Name
    Data.append(["Image 3","Detailed Product View 5"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 6")#Name
    Data.append(["Image 4","Detailed Product View 6"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 7")#Name
    Data.append(["Image 5","Detailed Product View 7"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 8")#Name
    Data.append(["Image 6","Detailed Product View 8"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 9")#Name
    Data.append(["Image 7","Detailed Product View 9"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 10")#Name
    Data.append(["Image 8","Detailed Product View 10"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 11")#Name
    Data.append(["Image 9","Detailed Product View 11"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 12")#Name
    Data.append(["Image 10","Detailed Product View 12"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 13")#Name
    Data.append(["Image 11","Detailed Product View 13"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 14")#Name
    Data.append(["Image 12","Detailed Product View 14"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pic 15")#Name
    Data.append(["Image 13","Detailed Product View 15"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Spin")#Name
    Data.append(["Spin 1","360 Degree Spin (Lowe's)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Specifications")#Name
    Data.append(["Spec Sheet","Dimensions Guide"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Number of Drawers")#Name
    Data.append(["Number of Drawers"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Number of Doors")#Name
    Data.append(["Number of Doors"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Number of Sinks")#Name
    Data.append(["Number of Sinks"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Sink")#Name
    Data.append(["Sink"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Sink Included")#Name
    Data.append(["Sink Included"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Sink Type")#Name
    Data.append(["Sink Type"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Maximum Sink Depth")#Name
    Data.append(["Maximum Sink Depth"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Sink Color")#Name
    Data.append(["Sink Color"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Top Width/Sink Count")#Name
    Data.append(["Top Width/Sink Count"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    return Mapping

def GetMappingData(Mapping,Name):
    for Map in Mapping:
        if Map[0] == Name:
            return Map

def LoadWorkBook(FileName):
    return openpyxl.load_workbook(FileName)

def GetSpreadSheetNames(Workbook):
    return Workbook.sheetnames

def GetSpreadSheetColumnNames(Spreadsheet,Row):
    ColumnNamesList = []
    for Int in range(1,Spreadsheet.max_column+1):
        ColumnNamesList.append(Spreadsheet[GetColLetter(Int)+str(Row)].value)
    return ColumnNamesList

def GetColLetter(Col):
    return openpyxl.utils.get_column_letter(Col)

def GenerateProductSheet(Name,Type,Sheet,Row):
    return [Name,Type,Sheet,Row,GenerateMapping()]

def MapSheet(Name,Sheet,Row,Mapping):
    ColumnNamesList = GetSpreadSheetColumnNames(Sheet,Row)
    Location = 1
    for Entry in ColumnNamesList:
        for Map in Mapping:
            for Data in Map[1]:
                if Entry == Data:
                    print("Found '"+Map[0]+"' at "+str(Location)+" in '"+Name+"'")
                    Map[2] = Location
        Location += 1
    return Mapping

def Map(ProductSheet):
    for Sheet in ProductSheet:
        Sheet[4] = MapSheet(Sheet[0],Sheet[2],Sheet[3],Sheet[4])
    return ProductSheet

def CheckType(Input,Output,Location,Row):
    InputType = Input[1]
    OutputType = Output[1]
    Type = InputType
    if Type == 'Mix':
        ProductTypeColNum = GetMappingData(Input[4],'Product Type')
        CounterColNum = GetMappingData(Input[4],'Counter')
        ProductType = Input[2][GetColLetter(ProductTypeColNum[2])+str(Row+Input[3])].value
        Counter = Input[2][GetColLetter(CounterColNum[2])+str(Row+Input[3])].value
        if ProductType == 'Vanity Set':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Base Only':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity with Mirror no Counter':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity with Counter No Mirror':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Set with Side cabinet':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Set with Side cabinets':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Side Cabinet':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Make-up table':
            Type = 'Makeup Vanities'
        else:
            print('Error Unknown Product Type: '+ProductType)
        if Type == 'Bathroom Vanity':
            if Counter == 'Yes':
                Type = 'Bathroom Vanities with Tops'
            elif Counter == 'No':
                Type = 'Bathroom Vanities without Tops'
            else:
                print('No Counter Column Input for Bathroom Vanity')
    if Type == OutputType or OutputType == 'Mix':
        return True
    else:
        return False

def WriteCopy(Input,Output,Location):
    #InputName = Input[0]
    InputSheet = Input[2]
    InputSheetOffset = Input[3]
    InputMapping = Input[4]
    #OutputName = Output[0]
    OutputSheet = Output[2]
    OutputSheetOffset = Output[3]
    OutputMapping = Output[4]
    for Row in range(1,InputSheet.max_row+1-InputSheetOffset):
        if CheckType(Input,Output,Location,Row):
            Wrote = False
            InputSheetValue = InputSheet[GetColLetter(InputMapping[Location][2])+str(Row+InputSheetOffset)].value
            for WriteRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                OutputSheetValue = OutputSheet[GetColLetter(OutputMapping[Location][2])+str(WriteRow)].value
                if OutputSheetValue == None and Wrote == False:
                    OutputSheet[GetColLetter(OutputMapping[Location][2])+str(WriteRow)].value = str(InputSheetValue)
                    Wrote = True
                    break
            if Wrote == False:
                OutputSheetValue = OutputSheet[GetColLetter(OutputMapping[Location][2])+str(OutputSheet.max_row)].value
                if OutputSheetValue != None:
                    OutputSheetValue = OutputSheet[GetColLetter(OutputMapping[Location][2])+str(OutputSheet.max_row+1)].value
                    OutputSheet[GetColLetter(OutputMapping[Location][2])+str(OutputSheet.max_row)].value = str(InputSheetValue)
            #print(InputName,InputMapping[Location][0],InputSheetValue,OutputName,OutputMapping[Location][0],OutputSheetValue)

def AutoFill(Inputs,Outputs):
    for Input in Inputs:
        for Output in Outputs:
            if Input[1] == 'Mix' or Input[1] == Output[1]:
                Location = 0
                for Map in Input[4]:
                    if Map[2] != None and Output[4][Location][2] != None and Map[3] == True:
                        print("Copy '"+Input[0]+"' "+Map[0]+" ["+GetColLetter(Map[2])+"] to '"+Output[0]+"' "+Output[4][Location][0]+" ["+GetColLetter(Output[4][Location][2])+"]")
                        WriteCopy(Input,Output,Location)
                    Location += 1

def WriteDefault(Output,Location):
    Name = Output[0]
    Sheet = Output[2]
    SheetOffset = Output[3]
    Mapping = Output[4]
    for Row in range(1,Sheet.max_row+1-SheetOffset):
        ExistingData = Sheet[GetColLetter(Mapping[Location][2])+str(Row+SheetOffset)].value
        WrittingData = Mapping[Location][4]
        if WrittingData != None:
            if ExistingData != None:
                Sheet[GetColLetter(Mapping[Location][2])+str(Row+SheetOffset)].value = WrittingData
                #print(Name+" "+Mapping[Location][0]+" Row: "+str(Row+SheetOffset)+" Overwrote Existing Data '"+str(ExistingData)+"' WrittingData '"+str(WrittingData)+"'")
            else:
                Sheet[GetColLetter(Mapping[Location][2])+str(Row+SheetOffset)].value = WrittingData
                #print(Name+" "+Mapping[Location][0]+" Row: "+str(Row+SheetOffset)+" Wrote '"+str(WrittingData)+"'")
        else:
            print(Name+" "+" Row: "+str(Row+SheetOffset)+" WrittingData is 'None'")

def AutoDefault(Outputs):
    for Output in Outputs:
        Location = 0
        for Map in Output[4]:
            if Map[2] != None and Map[4] != None:
                print("Defaulting '"+Output[0]+"' "+Map[0]+" ["+GetColLetter(Map[2])+"] to '"+str(Map[4])+"'")
                WriteDefault(Output,Location)
            Location += 1

def CalculateNetWeightUnpackaged(VanityWeightOnly,CounterWeightOnly,MirrorWeightOnly):
    return str(round(float(str(VanityWeightOnly).replace('N/A','0'))+float(str(CounterWeightOnly).replace('N/A','0'))+float(str(MirrorWeightOnly).replace('N/A','0')),2))

def GenerateNetWeightUnpackaged(Inputs,Outputs):
    for Input in Inputs:
        InputName = Input[0]
        InputSheet = Input[2]
        InputSheetOffset = Input[3]
        InputMapping = Input[4]
        for Output in Outputs:
            OutputName = Output[0]
            OutputSheet = Output[2]
            OutputSheetOffset = Output[3]
            OutputMapping = Output[4]
            NetWeightUnpackaged = GetMappingData(OutputMapping,'Net Weight Unpackaged')
            ItemNetWeight = GetMappingData(OutputMapping,'Item Net Weight')
            Weight = GetMappingData(OutputMapping,'Weight')
            OutputUPC = GetMappingData(OutputMapping,'UPC')
            InputUPC = GetMappingData(InputMapping,'UPC')
            VanityWeightOnly = GetMappingData(InputMapping,'Vanity Weight Only')
            CounterWeightOnly = GetMappingData(InputMapping,'Counter Weight Only')
            MirrorWeightOnly = GetMappingData(InputMapping,'Mirror Weight Only')
            if NetWeightUnpackaged != None and OutputUPC != None and InputUPC != None and VanityWeightOnly != None and CounterWeightOnly != None and MirrorWeightOnly != None:
                Msg = "Generating '"+OutputName+"' "+NetWeightUnpackaged[0]+" ["+GetColLetter(NetWeightUnpackaged[2])+"] using '"+InputName+"' "+VanityWeightOnly[0]+" ["+GetColLetter(VanityWeightOnly[2])+"] & "
                Msg += CounterWeightOnly[0]+" ["+GetColLetter(CounterWeightOnly[2])+"] & "+MirrorWeightOnly[0]+" ["+GetColLetter(MirrorWeightOnly[2])+"]"
                print(Msg)
                for OutputRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                    OutputUPCValue = OutputSheet[GetColLetter(OutputUPC[2])+str(OutputRow)].value
                    Wrote = False
                    for InputRow in range(1+InputSheetOffset,InputSheet.max_row+1):
                        InputUPCValue = str(InputSheet[GetColLetter(InputUPC[2])+str(InputRow)].value)
                        if OutputUPCValue != None and InputUPCValue != None and str(OutputUPCValue) == str(InputUPCValue):
                            OutputNetWeightUnpackaged = OutputSheet[GetColLetter(NetWeightUnpackaged[2])+str(OutputRow)].value
                            ItemNetWeightValue = None
                            WeightValue = None
                            InputVanityWeightOnly = InputSheet[GetColLetter(VanityWeightOnly[2])+str(InputRow)].value
                            InputCounterWeightOnly = InputSheet[GetColLetter(CounterWeightOnly[2])+str(InputRow)].value
                            InputMirrorWeightOnly = InputSheet[GetColLetter(MirrorWeightOnly[2])+str(InputRow)].value
                            if ItemNetWeight[2] != None:
                                ItemNetWeightValue =  OutputSheet[GetColLetter(ItemNetWeight[2])+str(OutputRow)].value
                            if Weight[2] != None:
                                WeightValue = OutputSheet[GetColLetter(Weight[2])+str(OutputRow)].value
                            if InputVanityWeightOnly != None and InputCounterWeightOnly != None and InputMirrorWeightOnly != None:
                                CalculatedNetWeightUnpackaged = CalculateNetWeightUnpackaged(InputVanityWeightOnly,InputCounterWeightOnly,InputMirrorWeightOnly)
                            else:
                                print("Error '"+OutputName+"' Row: "+str(OutputRow)+" Can't Calculate "+NetWeightUnpackaged[0]+" ["+GetColLetter(NetWeightUnpackaged[2])+"] "+"of '"+InputName+"' Row: "+str(InputRow))
                                CalculatedNetWeightUnpackaged = 'Missing'
                            if OutputNetWeightUnpackaged != None:
                                OutputSheet[GetColLetter(NetWeightUnpackaged[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Net Weight Unpackaged of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(OutputNetWeightUnpackaged)+" New: "+CalculatedNetWeightUnpackaged)
                                if ItemNetWeight[2] != None:
                                    if ItemNetWeightValue != None:
                                        OutputSheet[GetColLetter(ItemNetWeight[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                        #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Item Net Weight of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(ItemNetWeightValue)+" New: "+CalculatedNetWeightUnpackaged)
                                    else:
                                        OutputSheet[GetColLetter(ItemNetWeight[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                        #print(OutputName+" Row: "+str(OutputRow)+" Wrote Item Net Weight of UPC: "+str(OutputUPCValue)+" to "+CalculatedNetWeightUnpackaged)
                                if Weight[2] != None:
                                    if WeightValue != None:
                                        OutputSheet[GetColLetter(Weight[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                        #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Weight of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(WeightValue)+" New: "+CalculatedNetWeightUnpackaged)
                                    else:
                                        OutputSheet[GetColLetter(Weight[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                        #print(OutputName+" Row: "+str(OutputRow)+" Wrote Weight of UPC: "+str(OutputUPCValue)+" to "+CalculatedNetWeightUnpackaged)
                                Wrote = True
                                break
                            else:
                                OutputSheet[GetColLetter(NetWeightUnpackaged[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                #print(OutputName+" Row: "+str(OutputRow)+" Wrote Net Weight Unpackaged of UPC: "+str(OutputUPCValue)+" to "+CalculatedNetWeightUnpackaged)
                                if ItemNetWeight[2] != None:
                                    if ItemNetWeightValue != None:
                                        OutputSheet[GetColLetter(ItemNetWeight[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                        #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Item Net Weight of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(ItemNetWeightValue)+" New: "+CalculatedNetWeightUnpackaged)
                                    else:
                                        OutputSheet[GetColLetter(ItemNetWeight[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                        #print(OutputName+" Row: "+str(OutputRow)+" Wrote Item Net Weight of UPC: "+str(OutputUPCValue)+" to "+CalculatedNetWeightUnpackaged)
                                if Weight[2] != None:
                                    if WeightValue != None:
                                        OutputSheet[GetColLetter(Weight[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                        #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Weight of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(WeightValue)+" New: "+CalculatedNetWeightUnpackaged)
                                    else:
                                        OutputSheet[GetColLetter(Weight[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                        #print(OutputName+" Row: "+str(OutputRow)+" Wrote Weight of UPC: "+str(OutputUPCValue)+" to "+CalculatedNetWeightUnpackaged)
                                Wrote = True
                                break
                        else:
                            if OutputUPCValue == None:
                                print("Error '"+OutputName+"' Row: "+str(OutputRow)+" No UPC")
                                break
                    if Wrote == False:
                        print("Error '"+OutputName+"' Unable to find UPC: "+str(OutputSheet[GetColLetter(OutputUPC[2])+str(OutputRow)].value))

def GenerateGTIN(Outputs):
    for Output in Outputs:
        OutputName = Output[0]
        OutputSheet = Output[2]
        OutputSheetOffset = Output[3]
        OutputMapping = Output[4]
        UPC = GetMappingData(OutputMapping,'UPC')
        GTIN = GetMappingData(OutputMapping,'GTIN')
        HighestLevelGTIN = GetMappingData(OutputMapping,'Highest Level GTIN')
        if UPC != None and UPC[2] != None and GTIN and GTIN[2] != None and HighestLevelGTIN and HighestLevelGTIN[2] != None:
            print("Generating '"+OutputName+"' "+GTIN[0])
            for OutputRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                OutputUPC = str(OutputSheet[GetColLetter(UPC[2])+str(OutputRow)].value)
                OutputGTIN = OutputSheet[GetColLetter(GTIN[2])+str(OutputRow)].value
                OutputHighestLevelGTIN = OutputSheet[GetColLetter(HighestLevelGTIN[2])+str(OutputRow)].value
                if OutputGTIN != None:
                    OutputSheet[GetColLetter(GTIN[2])+str(OutputRow)].value = '00'+OutputUPC
                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote GTIN of UPC: "+str(OutputUPC)+" to "+"Old: "+str(OutputGTIN)+" New: "+'00'+OutputUPC)
                else:
                    OutputSheet[GetColLetter(GTIN[2])+str(OutputRow)].value = '00'+OutputUPC
                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote GTIN of UPC: "+str(OutputUPC)+" to "+'00'+OutputUPC)
                if OutputHighestLevelGTIN != None:
                    OutputSheet[GetColLetter(HighestLevelGTIN[2])+str(OutputRow)].value = '00'+OutputUPC
                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Highest Level GTIN of UPC: "+str(OutputUPC)+" to "+"Old: "+str(OutputHighestLevelGTIN)+" New: "+'00'+OutputUPC)
                else:
                    OutputSheet[GetColLetter(HighestLevelGTIN[2])+str(OutputRow)].value = '00'+OutputUPC
                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote Highest Level GTIN of UPC: "+str(OutputUPC)+" to "+'00'+OutputUPC)

def GenerateBullets(Inputs,Outputs):
    for Input in Inputs:
        InputName = Input[0]
        InputSheet = Input[2]
        InputSheetOffset = Input[3]
        InputMapping = Input[4]
        for Output in Outputs:
            OutputName = Output[0]
            OutputSheet = Output[2]
            OutputSheetOffset = Output[3]
            OutputMapping = Output[4]
            OutputUPC = GetMappingData(OutputMapping,'UPC')
            InputUPC = GetMappingData(InputMapping,'UPC')
            OutputBP3 = GetMappingData(OutputMapping,'Bullet Point 3')
            OutputBP4 = GetMappingData(OutputMapping,'Bullet Point 4')
            OutputBP5 = GetMappingData(OutputMapping,'Bullet Point 5')
            OutputBP6 = GetMappingData(OutputMapping,'Bullet Point 6')
            OutputBP7 = GetMappingData(OutputMapping,'Bullet Point 7')
            OutputBP8 = GetMappingData(OutputMapping,'Bullet Point 8')
            OutputBP9 = GetMappingData(OutputMapping,'Bullet Point 9')
            InputBP3 = GetMappingData(InputMapping,'Bullet Point 3')
            InputBP4 = GetMappingData(InputMapping,'Bullet Point 4')
            InputBP5 = GetMappingData(InputMapping,'Bullet Point 5')
            InputBP6 = GetMappingData(InputMapping,'Bullet Point 6')
            InputBP7 = GetMappingData(InputMapping,'Bullet Point 7')
            InputBP8 = GetMappingData(InputMapping,'Bullet Point 8')
            InputBP9 = GetMappingData(InputMapping,'Bullet Point 9')
            InputBP10 = GetMappingData(InputMapping,'Bullet Point 10')
            InputBP11 = GetMappingData(InputMapping,'Bullet Point 11')
            InputBP12 = GetMappingData(InputMapping,'Bullet Point 12')
            InputBP13 = GetMappingData(InputMapping,'Bullet Point 13')
            InputBP14 = GetMappingData(InputMapping,'Bullet Point 14')
            InputBP15 = GetMappingData(InputMapping,'Bullet Point 15')
            InputBP16 = GetMappingData(InputMapping,'Bullet Point 16')
            if OutputBP3 != None and OutputBP4 != None and OutputBP5 != None and OutputBP6 != None and OutputBP7 != None and OutputBP8 != None and OutputBP9 != None and InputBP3 != None and InputBP4 != None and InputBP5 != None:
                if InputBP6 != None and InputBP7 != None and InputBP8 != None and InputBP9 != None and InputBP10 != None and InputBP11 != None and InputBP12 != None and InputBP13 != None and InputBP14 != None and InputBP15 != None and InputBP16 != None:
                    print("Generating '"+OutputName+"' Bullet Points using '"+InputName+"'")
                    for OutputRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                        OutputUPCValue = OutputSheet[GetColLetter(OutputUPC[2])+str(OutputRow)].value
                        for InputRow in range(1+InputSheetOffset,InputSheet.max_row+1):
                            InputUPCValue = str(InputSheet[GetColLetter(InputUPC[2])+str(InputRow)].value)
                            if OutputUPCValue != None and InputUPCValue != None and str(OutputUPCValue) == str(InputUPCValue):
                                OutputBP3Value = OutputSheet[GetColLetter(OutputBP3[2])+str(OutputRow)].value
                                OutputBP4Value = OutputSheet[GetColLetter(OutputBP4[2])+str(OutputRow)].value
                                OutputBP5Value = OutputSheet[GetColLetter(OutputBP5[2])+str(OutputRow)].value
                                OutputBP6Value = OutputSheet[GetColLetter(OutputBP6[2])+str(OutputRow)].value
                                OutputBP7Value = OutputSheet[GetColLetter(OutputBP7[2])+str(OutputRow)].value
                                OutputBP8Value = OutputSheet[GetColLetter(OutputBP8[2])+str(OutputRow)].value
                                OutputBP9Value = OutputSheet[GetColLetter(OutputBP9[2])+str(OutputRow)].value
                                InputBP3Value = InputSheet[GetColLetter(InputBP3[2])+str(InputRow)].value
                                InputBP4Value = InputSheet[GetColLetter(InputBP4[2])+str(InputRow)].value
                                InputBP5Value = InputSheet[GetColLetter(InputBP5[2])+str(InputRow)].value
                                InputBP6Value = InputSheet[GetColLetter(InputBP6[2])+str(InputRow)].value
                                InputBP7Value = InputSheet[GetColLetter(InputBP7[2])+str(InputRow)].value
                                InputBP8Value = InputSheet[GetColLetter(InputBP8[2])+str(InputRow)].value
                                InputBP9Value = InputSheet[GetColLetter(InputBP9[2])+str(InputRow)].value
                                InputBP10Value = InputSheet[GetColLetter(InputBP10[2])+str(InputRow)].value
                                InputBP11Value = InputSheet[GetColLetter(InputBP11[2])+str(InputRow)].value
                                InputBP12Value = InputSheet[GetColLetter(InputBP12[2])+str(InputRow)].value
                                InputBP13Value = InputSheet[GetColLetter(InputBP13[2])+str(InputRow)].value
                                InputBP14Value = InputSheet[GetColLetter(InputBP14[2])+str(InputRow)].value
                                InputBP15Value = InputSheet[GetColLetter(InputBP15[2])+str(InputRow)].value
                                InputBP16Value = InputSheet[GetColLetter(InputBP16[2])+str(InputRow)].value
                                InputBPS = [InputBP3Value,InputBP4Value,InputBP5Value,InputBP6Value,InputBP7Value,InputBP8Value,InputBP9Value,InputBP10Value,InputBP11Value,InputBP12Value,InputBP13Value,InputBP14Value,InputBP15Value,InputBP16Value]
                                for Index in range(0,len(InputBPS)):
                                    if None not in InputBPS:
                                        if 'N/A' in InputBPS:
                                            InputBPS.remove('N/A')
                                            InputBPS.append('')
                                        elif '#N/A' in InputBPS:
                                            InputBPS.remove('#N/A')
                                            InputBPS.append('')
                                        else:
                                            InputBPS[Index] = InputBPS[Index].strip()
                                    else:
                                        print("Error '"+OutputName+"' Row: "+str(OutputRow)+" Can't Get Bullet Point "+str(Index+3)+" of '"+InputName+"' Row: "+str(InputRow))
                                Concatenation = ''
                                for Index in range(6,len(InputBPS)):
                                    if InputBPS[Index] == '':
                                        continue
                                    else:
                                        if Concatenation == '':
                                            Concatenation = InputBPS[Index]
                                        else:
                                            Concatenation += ', '+InputBPS[Index]
                                InputBPS[6] = Concatenation
                                if OutputBP3Value != None:
                                    OutputSheet[GetColLetter(OutputBP3[2])+str(OutputRow)].value = InputBPS[0]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Bullet Point 3 of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(InputBPS[0])+" New: "+str(InputBPS[0]))
                                else:
                                    OutputSheet[GetColLetter(OutputBP3[2])+str(OutputRow)].value = InputBPS[0]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote Bullet Point 3 of UPC: "+str(OutputUPCValue)+" to "+InputBPS[0])
                                if OutputBP4Value != None:
                                    OutputSheet[GetColLetter(OutputBP4[2])+str(OutputRow)].value = InputBPS[1]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Bullet Point 4 of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(InputBPS[1])+" New: "+str(InputBPS[1]))
                                else:
                                    OutputSheet[GetColLetter(OutputBP4[2])+str(OutputRow)].value = InputBPS[1]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote Bullet Point 4 of UPC: "+str(OutputUPCValue)+" to "+InputBPS[1])
                                if OutputBP5Value != None:
                                    OutputSheet[GetColLetter(OutputBP5[2])+str(OutputRow)].value = InputBPS[2]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Bullet Point 5 of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(InputBPS[2])+" New: "+str(InputBPS[2]))
                                else:
                                    OutputSheet[GetColLetter(OutputBP5[2])+str(OutputRow)].value = InputBPS[2]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote Bullet Point 5 of UPC: "+str(OutputUPCValue)+" to "+InputBPS[2])
                                if OutputBP6Value != None:
                                    OutputSheet[GetColLetter(OutputBP6[2])+str(OutputRow)].value = InputBPS[3]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Bullet Point 6 of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(InputBPS[3])+" New: "+str(InputBPS[3]))
                                else:
                                    OutputSheet[GetColLetter(OutputBP6[2])+str(OutputRow)].value = InputBPS[3]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote Bullet Point 6 of UPC: "+str(OutputUPCValue)+" to "+InputBPS[3])
                                if OutputBP7Value != None:
                                    OutputSheet[GetColLetter(OutputBP7[2])+str(OutputRow)].value = InputBPS[4]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Bullet Point 7 of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(InputBPS[4])+" New: "+str(InputBPS[4]))
                                else:
                                    OutputSheet[GetColLetter(OutputBP7[2])+str(OutputRow)].value = InputBPS[4]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote Bullet Point 7 of UPC: "+str(OutputUPCValue)+" to "+InputBPS[4])
                                if OutputBP8Value != None:
                                    OutputSheet[GetColLetter(OutputBP8[2])+str(OutputRow)].value = InputBPS[5]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Bullet Point 8 of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(InputBPS[5])+" New: "+str(InputBPS[5]))
                                else:
                                    OutputSheet[GetColLetter(OutputBP8[2])+str(OutputRow)].value = InputBPS[5]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote Bullet Point 8 of UPC: "+str(OutputUPCValue)+" to "+InputBPS[5])
                                if OutputBP9Value != None:
                                    OutputSheet[GetColLetter(OutputBP9[2])+str(OutputRow)].value = InputBPS[6]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Bullet Point 9 of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(InputBPS[6])+" New: "+str(InputBPS[6]))
                                else:
                                    OutputSheet[GetColLetter(OutputBP9[2])+str(OutputRow)].value = InputBPS[6]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote Bullet Point 9 of UPC: "+str(OutputUPCValue)+" to "+InputBPS[6])
                            else:
                                if OutputUPCValue == None:
                                    print("Error '"+OutputName+"' Row: "+str(OutputRow)+" No UPC")
                                    break
                else:
                    print("Error '"+OutputName+" Missing Bullet Points using '"+InputName+"'")
            else:
                print("Error '"+OutputName+" Missing Bullet Points using '"+InputName+"'")

def GenerateDimensions(Inputs,Outputs):
    for Input in Inputs:
        InputName = Input[0]
        InputSheet = Input[2]
        InputSheetOffset = Input[3]
        InputMapping = Input[4]
        for Output in Outputs:
            OutputName = Output[0]
            OutputSheet = Output[2]
            OutputSheetOffset = Output[3]
            OutputMapping = Output[4]
            OutputUPC = GetMappingData(OutputMapping,'UPC')
            InputUPC = GetMappingData(InputMapping,'UPC')
            CommonWidth = GetMappingData(OutputMapping,'Common Width')
            CommonHeight = GetMappingData(OutputMapping,'Common Height')
            CommonDepth = GetMappingData(OutputMapping,'Common Depth')
            ActualWidth = GetMappingData(OutputMapping,'Actual Width')
            ActualHeight = GetMappingData(OutputMapping,'Actual Height')
            ActualDepth = GetMappingData(OutputMapping,'Actual Depth')
            CabinetWidth = GetMappingData(OutputMapping,'Cabinet Width')
            CabinetHeight = GetMappingData(OutputMapping,'Cabinet Height')
            CabinetDepth = GetMappingData(OutputMapping,'Cabinet Depth')
            VanityWidth = GetMappingData(OutputMapping,'Vanity Width')
            CompatibleVanityTopWidth = GetMappingData(OutputMapping,'Compatible Vanity Top Width')
            CompatibleVanityTopDepth = GetMappingData(OutputMapping,'Compatible Vanity Top Depth')
            ProductDims = GetMappingData(InputMapping,'Product Dims')
            CabinetDims = GetMappingData(InputMapping,'Cabinet Dims')
            if CommonWidth != None and CommonHeight != None and CommonDepth != None and ActualWidth != None and ActualHeight != None and ActualDepth != None and CabinetWidth != None and CabinetHeight != None and CabinetDepth != None:
                if CompatibleVanityTopWidth != None and CompatibleVanityTopDepth != None and ProductDims != None and CabinetDims != None:
                    print("Generating '"+OutputName+"' Dimensions using '"+InputName+"'")
                    for OutputRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                        OutputUPCValue = OutputSheet[GetColLetter(OutputUPC[2])+str(OutputRow)].value
                        for InputRow in range(1+InputSheetOffset,InputSheet.max_row+1):
                            InputUPCValue = str(InputSheet[GetColLetter(InputUPC[2])+str(InputRow)].value)
                            if OutputUPCValue != None and InputUPCValue != None and str(OutputUPCValue) == str(InputUPCValue):
                                CommonWidthValue = OutputSheet[GetColLetter(CommonWidth[2])+str(OutputRow)].value
                                CommonHeightValue = OutputSheet[GetColLetter(CommonHeight[2])+str(OutputRow)].value
                                CommonDepthValue = OutputSheet[GetColLetter(CommonDepth[2])+str(OutputRow)].value
                                ActualWidthValue = OutputSheet[GetColLetter(ActualWidth[2])+str(OutputRow)].value
                                ActualHeightValue = OutputSheet[GetColLetter(ActualHeight[2])+str(OutputRow)].value
                                ActualDepthValue = OutputSheet[GetColLetter(ActualDepth[2])+str(OutputRow)].value
                                CabinetWidthValue = OutputSheet[GetColLetter(CabinetWidth[2])+str(OutputRow)].value
                                CabinetHeightValue = OutputSheet[GetColLetter(CabinetHeight[2])+str(OutputRow)].value
                                CabinetDepthValue = OutputSheet[GetColLetter(CabinetDepth[2])+str(OutputRow)].value
                                HasVanityWidth = False
                                HasCompatibleVanityTop = False
                                if VanityWidth[2] != None:
                                    VanityWidthValue = OutputSheet[GetColLetter(VanityWidth[2])+str(OutputRow)].value
                                    HasVanityWidth = True
                                if CompatibleVanityTopWidth[2] != None and CompatibleVanityTopDepth[2] != None:
                                    CompatibleVanityTopWidthValue = OutputSheet[GetColLetter(CompatibleVanityTopWidth[2])+str(OutputRow)].value
                                    CompatibleVanityTopDepthValue = OutputSheet[GetColLetter(CompatibleVanityTopDepth[2])+str(OutputRow)].value
                                    HasCompatibleVanityTop = True
                                ProductDimsValue = InputSheet[GetColLetter(ProductDims[2])+str(InputRow)].value
                                CabinetDimsValue = InputSheet[GetColLetter(CabinetDims[2])+str(InputRow)].value
                                if ProductDimsValue == None or CabinetDimsValue == None:
                                    print("Error '"+OutputName+"' Row: "+str(OutputRow)+" Can't Get Dims "+str(Index+3)+" of '"+InputName+"' Row: "+str(InputRow))
                                    continue
                                ProductDimsValueList = ProductDimsValue.replace('"','').lower().split("x")
                                CabinetDimsValueList = CabinetDimsValue.replace('"','').lower().split("x")
                                for Index in range(0,len(ProductDimsValueList)):
                                    if '/' in ProductDimsValueList[Index] and ' ' in ProductDimsValueList[Index]:
                                        ProductDimsValueList[Index] = float(ProductDimsValueList[Index].split(" ")[0])+float(eval(ProductDimsValueList[Index].split(" ")[1]))
                                    elif '/' in ProductDimsValueList[Index] and '-' in ProductDimsValueList[Index]:
                                        ProductDimsValueList[Index] = float(ProductDimsValueList[Index].split("-")[0])+float(eval(ProductDimsValueList[Index].split("-")[1]))
                                for Index in range(0,len(CabinetDimsValueList)):
                                    if '/' in CabinetDimsValueList[Index] and ' ' in CabinetDimsValueList[Index]:
                                        CabinetDimsValueList[Index] = float(CabinetDimsValueList[Index].split(" ")[0])+float(eval(CabinetDimsValueList[Index].split(" ")[1]))
                                    elif '/' in CabinetDimsValueList[Index] and '-' in CabinetDimsValueList[Index]:
                                        CabinetDimsValueList[Index] = float(CabinetDimsValueList[Index].split("-")[0])+float(eval(CabinetDimsValueList[Index].split("-")[1]))
                                for Index in range(0,len(ProductDimsValueList)):
                                    ProductDimsValueList[Index] = str(ProductDimsValueList[Index])
                                for Index in range(0,len(CabinetDimsValueList)):
                                    CabinetDimsValueList[Index] = str(CabinetDimsValueList[Index])
                                if CommonWidthValue != None:
                                    OutputSheet[GetColLetter(CommonWidth[2])+str(OutputRow)].value = ProductDimsValueList[0]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+CommonWidth[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(CommonWidthValue)+" New: "+str(ProductDimsValueList[0]))
                                else:
                                    OutputSheet[GetColLetter(CommonWidth[2])+str(OutputRow)].value = ProductDimsValueList[0]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+CommonWidth[0]+" UPC: "+str(OutputUPCValue)+" to "+ProductDimsValueList[0])
                                if CommonHeightValue != None:
                                    OutputSheet[GetColLetter(CommonHeight[2])+str(OutputRow)].value = ProductDimsValueList[2]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+CommonHeight[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(CommonHeightValue)+" New: "+str(ProductDimsValueList[2]))
                                else:
                                    OutputSheet[GetColLetter(CommonHeight[2])+str(OutputRow)].value = ProductDimsValueList[2]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+CommonHeight[0]+" of UPC: "+str(OutputUPCValue)+" to "+ProductDimsValueList[2])
                                if CommonDepthValue != None:
                                    OutputSheet[GetColLetter(CommonDepth[2])+str(OutputRow)].value = ProductDimsValueList[1]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+CommonDepth[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(CommonDepthValue)+" New: "+str(ProductDimsValueList[1]))
                                else:
                                    OutputSheet[GetColLetter(CommonDepth[2])+str(OutputRow)].value = ProductDimsValueList[1]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+CommonDepth[0]+" of UPC: "+str(OutputUPCValue)+" to "+ProductDimsValueList[1])
                                if ActualWidthValue != None:
                                    OutputSheet[GetColLetter(ActualWidth[2])+str(OutputRow)].value = ProductDimsValueList[0]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+ActualWidth[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(ActualWidthValue)+" New: "+str(ProductDimsValueList[0]))
                                else:
                                    OutputSheet[GetColLetter(ActualWidth[2])+str(OutputRow)].value = ProductDimsValueList[0]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+ActualWidth[0]+" UPC: "+str(OutputUPCValue)+" to "+ProductDimsValueList[0])
                                if ActualHeightValue != None:
                                    OutputSheet[GetColLetter(ActualHeight[2])+str(OutputRow)].value = ProductDimsValueList[2]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+ActualHeight[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(ActualHeightValue)+" New: "+str(ProductDimsValueList[2]))
                                else:
                                    OutputSheet[GetColLetter(ActualHeight[2])+str(OutputRow)].value = ProductDimsValueList[2]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+ActualHeight[0]+" of UPC: "+str(OutputUPCValue)+" to "+ProductDimsValueList[2])
                                if ActualDepthValue != None:
                                    OutputSheet[GetColLetter(ActualDepth[2])+str(OutputRow)].value = ProductDimsValueList[1]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+ActualDepth[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(ActualDepthValue)+" New: "+str(ProductDimsValueList[1]))
                                else:
                                    OutputSheet[GetColLetter(ActualDepth[2])+str(OutputRow)].value = ProductDimsValueList[1]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+ActualDepth[0]+" of UPC: "+str(OutputUPCValue)+" to "+ProductDimsValueList[1])
                                if CabinetWidthValue != None:
                                    OutputSheet[GetColLetter(CabinetWidth[2])+str(OutputRow)].value = CabinetDimsValueList[0]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+CabinetWidth[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(CabinetWidthValue)+" New: "+str(CabinetDimsValueList[0]))
                                else:
                                    OutputSheet[GetColLetter(CabinetWidth[2])+str(OutputRow)].value = CabinetDimsValueList[0]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+CabinetWidth[0]+" UPC: "+str(OutputUPCValue)+" to "+CabinetDimsValueList[0])
                                if CabinetHeightValue != None:
                                    OutputSheet[GetColLetter(CabinetHeight[2])+str(OutputRow)].value = CabinetDimsValueList[2]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+CabinetHeight[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(CabinetHeightValue)+" New: "+str(CabinetDimsValueList[2]))
                                else:
                                    OutputSheet[GetColLetter(CabinetHeight[2])+str(OutputRow)].value = CabinetDimsValueList[2]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+CabinetHeight[0]+" of UPC: "+str(OutputUPCValue)+" to "+CabinetDimsValueList[2])
                                if CabinetDepthValue != None:
                                    OutputSheet[GetColLetter(CabinetDepth[2])+str(OutputRow)].value = CabinetDimsValueList[1]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+CabinetDepth[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(CabinetDepthValue)+" New: "+str(CabinetDimsValueList[1]))
                                else:
                                    OutputSheet[GetColLetter(CabinetDepth[2])+str(OutputRow)].value = CabinetDimsValueList[1]
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+CabinetDepth[0]+" of UPC: "+str(OutputUPCValue)+" to "+CabinetDimsValueList[1])
                                if HasVanityWidth == True:
                                    if VanityWidthValue != None:
                                        OutputSheet[GetColLetter(VanityWidth[2])+str(OutputRow)].value = ProductDimsValueList[0]
                                        #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+VanityWidth[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(VanityWidthValue)+" New: "+str(ProductDimsValueList[0]))
                                    else:
                                        OutputSheet[GetColLetter(VanityWidth[2])+str(OutputRow)].value = ProductDimsValueList[0]
                                        #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+VanityWidth[0]+" of UPC: "+str(OutputUPCValue)+" to "+ProductDimsValueList[0])
                                if HasCompatibleVanityTop == True:
                                    if CompatibleVanityTopWidthValue != None:
                                        OutputSheet[GetColLetter(CompatibleVanityTopWidth[2])+str(OutputRow)].value = str(float(CabinetDimsValueList[0])+1).replace('.0','')
                                        #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+CompatibleVanityTopWidth[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(CompatibleVanityTopWidthValue)+" New: "+str(float(CabinetDimsValueList[0])+1).replace('.0',''))
                                    else:
                                        OutputSheet[GetColLetter(CompatibleVanityTopWidth[2])+str(OutputRow)].value = str(float(CabinetDimsValueList[0])+1).replace('.0','')
                                        #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+CompatibleVanityTopWidth[0]+" of UPC: "+str(OutputUPCValue)+" to "+str(float(CabinetDimsValueList[0])+1).replace('.0',''))
                                    if CompatibleVanityTopDepthValue != None:
                                        OutputSheet[GetColLetter(CompatibleVanityTopDepth[2])+str(OutputRow)].value = str(float(CabinetDimsValueList[1])+1).replace('.0','')
                                        #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+CompatibleVanityTopDepth[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(CompatibleVanityTopDepthValue)+" New: "+str(float(CabinetDimsValueList[1])+1).replace('.0',''))
                                    else:
                                        OutputSheet[GetColLetter(CompatibleVanityTopDepth[2])+str(OutputRow)].value = str(float(CabinetDimsValueList[1])+1).replace('.0','')
                                        #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+CompatibleVanityTopDepth[0]+" of UPC: "+str(OutputUPCValue)+" to "+str(float(CabinetDimsValueList[1])+1).replace('.0',''))
                            else:
                                if OutputUPCValue == None:
                                    print("Error '"+OutputName+"' Row: "+str(OutputRow)+" No UPC")
                                    break
                else:
                    print("Error '"+OutputName+" Missing Dimensions using '"+InputName+"'")
            else:
                print("Error '"+OutputName+" Missing Dimensions using '"+InputName+"'")

def GetFilename(Link):
    return str(Link.split('/')[len(Link.split('/'))-1])

def GenerateAssets(Inputs,Outputs):
    for Input in Inputs:
        InputName = Input[0]
        InputSheet = Input[2]
        InputSheetOffset = Input[3]
        InputMapping = Input[4]
        for Output in Outputs:
            OutputName = Output[0]
            OutputSheet = Output[2]
            OutputSheetOffset = Output[3]
            OutputMapping = Output[4]
            OutputUPC = GetMappingData(OutputMapping,'UPC')
            InputUPC = GetMappingData(InputMapping,'UPC')
            OutputPic1 = GetMappingData(OutputMapping,'Pic 1')
            OutputPic2 = GetMappingData(OutputMapping,'Pic 2')
            OutputPic3 = GetMappingData(OutputMapping,'Pic 3')
            OutputPic4 = GetMappingData(OutputMapping,'Pic 4')
            OutputPic5 = GetMappingData(OutputMapping,'Pic 5')
            OutputPic6 = GetMappingData(OutputMapping,'Pic 6')
            OutputPic7 = GetMappingData(OutputMapping,'Pic 7')
            OutputPic8 = GetMappingData(OutputMapping,'Pic 8')
            OutputPic9 = GetMappingData(OutputMapping,'Pic 9')
            OutputPic10 = GetMappingData(OutputMapping,'Pic 10')
            OutputPic11 = GetMappingData(OutputMapping,'Pic 11')
            OutputPic12 = GetMappingData(OutputMapping,'Pic 12')
            OutputPic13 = GetMappingData(OutputMapping,'Pic 13')
            OutputPic14 = GetMappingData(OutputMapping,'Pic 14')
            OutputPic15 = GetMappingData(OutputMapping,'Pic 15')
            OutputSpin = GetMappingData(OutputMapping,'Spin')
            OutputSpecifications = GetMappingData(OutputMapping,'Specifications')
            InputPic1 = GetMappingData(InputMapping,'Pic 1')
            InputPic2 = GetMappingData(InputMapping,'Pic 2')
            InputPic3 = GetMappingData(InputMapping,'Pic 3')
            InputPic4 = GetMappingData(InputMapping,'Pic 4')
            InputPic5 = GetMappingData(InputMapping,'Pic 5')
            InputPic6 = GetMappingData(InputMapping,'Pic 6')
            InputPic7 = GetMappingData(InputMapping,'Pic 7')
            InputPic8 = GetMappingData(InputMapping,'Pic 8')
            InputPic9 = GetMappingData(InputMapping,'Pic 9')
            InputPic10 = GetMappingData(InputMapping,'Pic 10')
            InputPic11 = GetMappingData(InputMapping,'Pic 11')
            InputPic12 = GetMappingData(InputMapping,'Pic 12')
            InputPic13 = GetMappingData(InputMapping,'Pic 13')
            InputPic14 = GetMappingData(InputMapping,'Pic 14')
            InputPic15 = GetMappingData(InputMapping,'Pic 15')
            InputSpin = GetMappingData(InputMapping,'Spin')
            InputSpecifications = GetMappingData(InputMapping,'Specifications')
            OutputPics = [OutputPic1,OutputPic2,OutputPic3,OutputPic4,OutputPic5,OutputPic6,OutputPic7,OutputPic8,OutputPic9,OutputPic10,OutputPic11,OutputPic12,OutputPic13,OutputPic14,OutputPic15]
            InputPics = [InputPic1,InputPic2,InputPic3,InputPic4,InputPic5,InputPic6,InputPic7,InputPic8,InputPic9,InputPic10,InputPic11,InputPic12,InputPic13,InputPic14,InputPic15]
            IOPics = OutputPics+InputPics+[OutputSpin,OutputSpecifications,InputSpin,InputSpecifications]
            print("Generating '"+OutputName+"' Assets using '"+InputName+"'")
            Break = False
            for IO in IOPics:
                if IO == None:
                    print("Error '"+OutputName+" Missing Asset using '"+InputName+"'")
                    Break = True
            if Break == True:
                Break = False
                break
            for OutputRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                OutputUPCValue = OutputSheet[GetColLetter(OutputUPC[2])+str(OutputRow)].value
                for InputRow in range(1+InputSheetOffset,InputSheet.max_row+1):
                    InputUPCValue = str(InputSheet[GetColLetter(InputUPC[2])+str(InputRow)].value)
                    if OutputUPCValue != None and InputUPCValue != None and str(OutputUPCValue) == str(InputUPCValue):
                        OutputPicsValue = []
                        InputPicsValue = []
                        HasSpin = False
                        HasSpecifications = False
                        for Index in range(0,len(OutputPics)):
                            OutputPicsValue.append(OutputSheet[GetColLetter(OutputPics[Index][2])+str(OutputRow)].value)
                        for Index in range(0,len(InputPics)):
                            if InputSheet[GetColLetter(InputPics[Index][2])+str(InputRow)].value != None:
                                InputPicsValue.append(GetFilename(InputSheet[GetColLetter(InputPics[Index][2])+str(InputRow)].value))
                        if InputSpin[2] != None and InputSheet[GetColLetter(InputSpin[2])+str(InputRow)].value != None:
                            InputSpinValue = GetFilename(InputSheet[GetColLetter(InputSpin[2])+str(InputRow)].value)
                            OutputSpinValue = OutputSheet[GetColLetter(OutputSpin[2])+str(OutputRow)].value
                            HasSpin = True
                        if InputSpecifications[2] != None and InputSheet[GetColLetter(InputSpecifications[2])+str(InputRow)].value != None:
                            InputSpecificationsValue = GetFilename(InputSheet[GetColLetter(InputSpecifications[2])+str(InputRow)].value)
                            OutputSpecificationsValue = OutputSheet[GetColLetter(OutputSpecifications[2])+str(OutputRow)].value
                            HasSpecifications = True
                        for Index in range(0,len(InputPicsValue)):
                            if OutputPicsValue[Index] != None:
                                OutputSheet[GetColLetter(OutputPics[Index][2])+str(OutputRow)].value = InputPicsValue[Index]
                                #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+OutputPics[Index][0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(OutputPicsValue[Index])+" New: "+str(InputPicsValue[Index]))
                            else:
                                OutputSheet[GetColLetter(OutputPics[Index][2])+str(OutputRow)].value = InputPicsValue[Index]
                                #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+OutputPics[Index][0]+" of UPC: "+str(OutputUPCValue)+" to "+str(InputPicsValue[Index]))
                        if HasSpin == True:
                            if OutputSpinValue != None:
                                OutputSheet[GetColLetter(OutputSpin[2])+str(OutputRow)].value = InputSpinValue
                                #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+OutputSpin[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+OutputSpinValue+" New: "+str(InputSpinValue))
                            else:
                                OutputSheet[GetColLetter(OutputSpin[2])+str(OutputRow)].value = InputSpinValue
                                #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+OutputSpin[0]+" of UPC: "+str(OutputUPCValue)+" to "+str(InputSpinValue))
                        if HasSpecifications == True:
                            if OutputSpecificationsValue != None:
                                OutputSheet[GetColLetter(OutputSpecifications[2])+str(OutputRow)].value = InputSpecificationsValue
                                #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+OutputSpecifications[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(OutputPicsValue[Index])+" New: "+str(InputSpecificationsValue))
                            else:
                                OutputSheet[GetColLetter(OutputSpecifications[2])+str(OutputRow)].value = InputSpecificationsValue
                                #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+OutputSpecifications[0]+" of UPC: "+str(OutputUPCValue)+" to "+str(InputSpecificationsValue))
                    else:
                        if OutputUPCValue == None:
                            print("Error '"+OutputName+"' Row: "+str(OutputRow)+" No UPC")
                            break

def ColorCode(Color):
    Code = Color.strip()
    if 'Blue' in Code:
        Code = 'Blue'
    elif 'Grey' in Code:
        Code = 'Gray'
    elif 'Brown' in Code:
        Code = 'Brown'
    elif 'Espresso' == Code:
        Code = 'Brown'
    return Code

def GenerateColorFamily(Outputs):
    for Output in Outputs:
        OutputName = Output[0]
        OutputSheet = Output[2]
        OutputSheetOffset = Output[3]
        OutputMapping = Output[4]
        UPC = GetMappingData(OutputMapping,'UPC')
        VanityColor = GetMappingData(OutputMapping,'Vanity Color')
        ColorFamily = GetMappingData(OutputMapping,'Color/Finish Family')
        if UPC != None and UPC[2] != None and VanityColor and VanityColor[2] != None and ColorFamily and ColorFamily[2] != None:
            print("Generating '"+OutputName+"' "+ColorFamily[0])
            for OutputRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                OutputUPC = str(OutputSheet[GetColLetter(UPC[2])+str(OutputRow)].value)
                OutputVanityColor = OutputSheet[GetColLetter(VanityColor[2])+str(OutputRow)].value
                OutputColorFamily = OutputSheet[GetColLetter(ColorFamily[2])+str(OutputRow)].value
                if OutputVanityColor != None:
                    ColorCodeValue = ColorCode(OutputVanityColor)
                else:
                    print("Error '"+OutputName+"' Row: "+str(OutputRow)+" No Vanity Color")
                    break
                if OutputColorFamily != None:
                    OutputSheet[GetColLetter(ColorFamily[2])+str(OutputRow)].value = ColorCodeValue
                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+ColorFamily+" of UPC: "+str(OutputUPC)+" to "+"Old: "+str(OutputColorFamily)+" New: "+str(ColorCodeValue))
                else:
                    OutputSheet[GetColLetter(ColorFamily[2])+str(OutputRow)].value = ColorCodeValue
                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+ColorFamily+" of UPC: "+str(OutputUPC)+" to "+str(ColorCodeValue))

def GenerateSinks(Inputs,Outputs):
    for Input in Inputs:
        InputName = Input[0]
        InputSheet = Input[2]
        InputSheetOffset = Input[3]
        InputMapping = Input[4]
        for Output in Outputs:
            OutputName = Output[0]
            OutputSheet = Output[2]
            OutputSheetOffset = Output[3]
            OutputMapping = Output[4]
            OutputUPC = GetMappingData(OutputMapping,'UPC')
            InputUPC = GetMappingData(InputMapping,'UPC')
            SinkIncluded = GetMappingData(OutputMapping,'Sink Included')
            SinkType = GetMappingData(OutputMapping,'Sink Type')
            SinkColor = GetMappingData(OutputMapping,'Sink Color')
            NumberofSinks = GetMappingData(OutputMapping,'Number of Sinks')
            MaximumSinkDepth = GetMappingData(OutputMapping,'Maximum Sink Depth')
            TopWidthSinkCount = GetMappingData(OutputMapping,'Top Width/Sink Count')
            Sink = GetMappingData(InputMapping,'Sink')
            SinkDims = GetMappingData(InputMapping,'Sink Dims')
            SinkBP = GetMappingData(InputMapping,'Bullet Point 12')
            OutputSinks = [SinkType,MaximumSinkDepth]
            InputSinks = [Sink,SinkDims,SinkBP]
            IOSinks = OutputSinks+InputSinks+[SinkIncluded,SinkColor,NumberofSinks+TopWidthSinkCount]
            print("Generating '"+OutputName+"' Sinks using '"+InputName+"'")
            Break = False
            for IO in IOSinks:
                if IO == None:
                    print("Error '"+OutputName+" Missing Sinks using '"+InputName+"'")
                    Break = True
            if Break == True:
                Break = False
                break
            for OutputRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                OutputUPCValue = OutputSheet[GetColLetter(OutputUPC[2])+str(OutputRow)].value
                for InputRow in range(1+InputSheetOffset,InputSheet.max_row+1):
                    InputUPCValue = str(InputSheet[GetColLetter(InputUPC[2])+str(InputRow)].value)
                    if OutputUPCValue != None and InputUPCValue != None and str(OutputUPCValue) == str(InputUPCValue):
                        OutputSinksValue = []
                        InputSinksValue = []
                        HasSinkIncluded = False
                        HasSinkColor = False
                        HasNumberofSinks = False
                        HasTopWidthSinkCount = False
                        for Index in range(0,len(OutputSinks)):
                            OutputSinksValue.append(OutputSheet[GetColLetter(OutputSinks[Index][2])+str(OutputRow)].value)
                        for Index in range(0,len(InputSinks)):
                            if InputSheet[GetColLetter(InputSinks[Index][2])+str(InputRow)].value != None:
                                InputSinksValue.append(InputSheet[GetColLetter(InputSinks[Index][2])+str(InputRow)].value)
                            else:
                                print("Error '"+OutputName+" Missing "+InputSinks[Index][0]+" using '"+InputName+"'")
                                Break = True
                        if Break == True:
                            Break = False
                            break
                        if SinkIncluded[2] != None:
                            SinkIncludedValue = OutputSheet[GetColLetter(SinkIncluded[2])+str(OutputRow)].value
                            HasSinkIncluded = True
                        if SinkColor[2] != None:
                            SinkColorValue = OutputSheet[GetColLetter(SinkColor[2])+str(OutputRow)].value
                            HasSinkColor = True
                        if NumberofSinks[2] != None:
                            NumberofSinksValue = OutputSheet[GetColLetter(NumberofSinks[2])+str(OutputRow)].value
                            HasNumberofSinks = True
                        if TopWidthSinkCount[2] != None:
                            TopWidthSinkCountValue = OutputSheet[GetColLetter(TopWidthSinkCount[2])+str(OutputRow)].value
                            HasTopWidthSinkCount = True
                        SinkDimsValueList = InputSinksValue[1].replace('"','').replace('”','').lower().split("x")
                        for Index in range(0,len(SinkDimsValueList)):
                            if '/' in SinkDimsValueList[Index] and ' ' in SinkDimsValueList[Index]:
                                SinkDimsValueList[Index] = str(float(SinkDimsValueList[Index].split(" ")[0])+float(eval(SinkDimsValueList[Index].split(" ")[1])))
                            elif '/' in SinkDimsValueList[Index] and '-' in SinkDimsValueList[Index]:
                                SinkDimsValueList[Index] = str(float(SinkDimsValueList[Index].split("-")[0])+float(eval(SinkDimsValueList[Index].split("-")[1])))
                        if 'single' in InputSinksValue[2].lower():
                            NumberofSinksString = 'Single sink'
                            TrueSinkWidth = SinkDimsValueList[0]
                        elif 'double' in InputSinksValue[2].lower():
                            NumberofSinksString = 'Double sink'
                            TrueSinkWidth = str(int(SinkDimsValueList[0])*2)
                        for Index in range(0,len(SinkDimsValueList)):
                            SinkDimsValueList[Index] = str(SinkDimsValueList[Index])
                        if HasSinkIncluded == True:
                            if InputSinksValue[0] !='N/A':
                                if SinkIncludedValue != None:
                                    OutputSheet[GetColLetter(SinkIncluded[2])+str(OutputRow)].value = 'Yes'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+SinkIncluded[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+SinkIncludedValue+" New: "+'Yes')
                                else:
                                    OutputSheet[GetColLetter(SinkIncluded[2])+str(OutputRow)].value = 'Yes'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+SinkIncluded[0]+" of UPC: "+str(OutputUPCValue)+" to "+'Yes')
                            else:
                                if SinkIncludedValue != None:
                                    OutputSheet[GetColLetter(SinkIncluded[2])+str(OutputRow)].value = 'No'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+SinkIncluded[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+SinkIncludedValue+" New: "+'No')
                                else:
                                    OutputSheet[GetColLetter(SinkIncluded[2])+str(OutputRow)].value = 'No'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+SinkIncluded[0]+" of UPC: "+str(OutputUPCValue)+" to "+'No')
                        if HasSinkColor == True:
                            if InputSinksValue[0] !='N/A':
                                if SinkColorValue != None:
                                    OutputSheet[GetColLetter(SinkColor[2])+str(OutputRow)].value = 'White'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+SinkColor[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+SinkColorValue+" New: "+'White')
                                else:
                                    OutputSheet[GetColLetter(SinkColor[2])+str(OutputRow)].value = 'White'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+SinkColor[0]+" of UPC: "+str(OutputUPCValue)+" to "+'White')
                            else:
                                if SinkColorValue != None:
                                    OutputSheet[GetColLetter(SinkColor[2])+str(OutputRow)].value = 'N/A'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+SinkColor[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+SinkColorValue+" New: "+'N/A')
                                else:
                                    OutputSheet[GetColLetter(SinkColor[2])+str(OutputRow)].value = 'N/A'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+SinkColor[0]+" of UPC: "+str(OutputUPCValue)+" to "+'N/A')
                        if HasNumberofSinks == True:
                            if InputSinksValue[0] !='N/A':
                                if NumberofSinksValue != None:
                                    OutputSheet[GetColLetter(NumberofSinks[2])+str(OutputRow)].value = NumberofSinksString
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+NumberofSinks[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+NumberofSinksValue+" New: "+NumberofSinksString)
                                else:
                                    OutputSheet[GetColLetter(NumberofSinks[2])+str(OutputRow)].value = NumberofSinksString
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+NumberofSinks[0]+" of UPC: "+str(OutputUPCValue)+" to "+NumberofSinksString)
                            else:
                                if NumberofSinksValue != None:
                                    OutputSheet[GetColLetter(NumberofSinks[2])+str(OutputRow)].value = 'N/A (no sink)'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "NumberofSinks[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+NumberofSinksValue+" New: "+'N/A (no sink)')
                                else:
                                    OutputSheet[GetColLetter(NumberofSinks[2])+str(OutputRow)].value = 'N/A (no sink)'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+NumberofSinks[0]+" of UPC: "+str(OutputUPCValue)+" to "+'N/A (no sink)')
                        if HasTopWidthSinkCount == True:
                            if InputSinksValue[0] !='N/A':
                                if TopWidthSinkCountValue != None:
                                    OutputSheet[GetColLetter(TopWidthSinkCount[2])+str(OutputRow)].value = TrueSinkWidth+'-in '+NumberofSinksString.lower()
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+TopWidthSinkCount[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+TopWidthSinkCountValue+" New: "+SinkDimsValueList[0]+'- 'TrueSinkWidth+'-in '+NumberofSinksString.lower())
                                else:
                                    OutputSheet[GetColLetter(TopWidthSinkCount[2])+str(OutputRow)].value = TrueSinkWidth+'-in '+NumberofSinksString.lower()
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+TopWidthSinkCount[0]+" of UPC: "+str(OutputUPCValue)+" to "+SinkDimsValueList[0]+'- '+TrueSinkWidth+'-in '+NumberofSinksString.lower())
                            else:
                                if TopWidthSinkCountValue != None:
                                    OutputSheet[GetColLetter(TopWidthSinkCount[2])+str(OutputRow)].value = 'N/A (no sink)'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "TopWidthSinkCount[0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+TopWidthSinkCountValue+" New: "+'N/A (no sink)')
                                else:
                                    OutputSheet[GetColLetter(TopWidthSinkCount[2])+str(OutputRow)].value = 'N/A (no sink)'
                                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+TopWidthSinkCount[0]+" of UPC: "+str(OutputUPCValue)+" to "+'N/A (no sink)')
                        if InputSinksValue[0] !='N/A':
                            if OutputSinksValue[0] != None:
                                OutputSheet[GetColLetter(OutputSinks[0][2])+str(OutputRow)].value = 'Undermount'
                                #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+OutputSinks[0][0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+OutputSinksValue[0]+" New: "+'Undermount')
                            else:
                                OutputSheet[GetColLetter(OutputSinks[0][2])+str(OutputRow)].value = 'Undermount'
                                #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+OutputSinks[0][0]+" of UPC: "+str(OutputUPCValue)+" to "+'Undermount')
                        else:
                            if OutputSinksValue[0] != None:
                                OutputSheet[GetColLetter(OutputSinks[0][2])+str(OutputRow)].value = 'None'
                                #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+OutputSinks[0][0]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+OutputSinksValue[0]+" New: "+'None')
                            else:
                                OutputSheet[GetColLetter(OutputSinks[0][2])+str(OutputRow)].value = 'None'
                                #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+OutputSinks[0][0]+" of UPC: "+str(OutputUPCValue)+" to "+'None')
                        if InputSinksValue[0] !='N/A':
                            if OutputSinksValue[1] != None:
                                OutputSheet[GetColLetter(OutputSinks[1][2])+str(OutputRow)].value = SinkDimsValueList[2]
                                #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+OutputSinks[1][1]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+OutputSinksValue[1]+" New: "+SinkDimsValueList[2])
                            else:
                                OutputSheet[GetColLetter(OutputSinks[1][2])+str(OutputRow)].value = SinkDimsValueList[2]
                                #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+OutputSinks[1][1]+" of UPC: "+str(OutputUPCValue)+" to "+SinkDimsValueList[2])
                        else:
                            if OutputSinksValue[1] != None:
                                OutputSheet[GetColLetter(OutputSinks[1][2])+str(OutputRow)].value = 'N/A'
                                #print(OutputName+" Row: "+str(OutputRow)+" Overwrote "+OutputSinks[1][1]+" of UPC: "+str(OutputUPCValue)+" to "+"Old: "+OutputSinksValue[1]+" New: "+'N/A')
                            else:
                                OutputSheet[GetColLetter(OutputSinks[1][2])+str(OutputRow)].value = 'N/A'
                                #print(OutputName+" Row: "+str(OutputRow)+" Wrote "+OutputSinks[1][1]+" of UPC: "+str(OutputUPCValue)+" to "+'N/A')
                    else:
                        if OutputUPCValue == None:
                            print("Error '"+OutputName+"' Row: "+str(OutputRow)+" No UPC")
                            break

def main():
    #Load Input Workbooks
    LexoraSourceWorkBook = LoadWorkBook("Lexora Source.xlsx")
    #Load Input Workbooks

    #Load Output Workbooks
    LowesBathroomVanitieswithTopsWorkBook = LoadWorkBook("Bathroom Vanities with Tops.xlsx")
    LowesBathroomVanitieswithoutTopsWorkBook = LoadWorkBook("Bathroom Vanities without Tops.xlsx")
    ###LowesMakeupVanitiesWorkBook = LoadWorkBook("Makeup Vanities.xlsx")
    #Load Output Workbooks

    #Load Input Sheets
    BathroomVanities = LexoraSourceWorkBook['Bathroom Vanities']
    #Load Input Sheets

    #Load Output Sheets
    BathroomVanitieswithTops = LowesBathroomVanitieswithTopsWorkBook['Attributes']
    BathroomVanitieswithoutTops = LowesBathroomVanitieswithoutTopsWorkBook['Attributes']
    ###MakeupVanities = LowesMakeupVanitiesWorkBook['Attributes']
    #Load Output Sheets

    #Load Input List
    Inputs = []
    Inputs.append(GenerateProductSheet('Bathroom Vanities','Mix',BathroomVanities,1))
    #Load Input List

    #Load Output List
    Outputs = []
    Outputs.append(GenerateProductSheet('Bathroom Vanities with Tops','Bathroom Vanities with Tops',BathroomVanitieswithTops,5))
    Outputs.append(GenerateProductSheet('Bathroom Vanities without Tops','Bathroom Vanities without Tops',BathroomVanitieswithoutTops,5))
    ###Outputs.append(GenerateProductSheet('Makeup Vanities','Makeup Vanities',MakeupVanities,5))
    #Load Output List

    #Mapping
    Inputs = Map(Inputs)
    Outputs = Map(Outputs)
    #Mapping

    #AutoFill
    AutoFill(Inputs,Outputs)
    #AutoFill

    #AutoDefault
    AutoDefault(Outputs)
    #AutoDefault

    #Generators
    GenerateNetWeightUnpackaged(Inputs,Outputs)
    GenerateGTIN(Outputs)
    GenerateBullets(Inputs,Outputs)
    GenerateDimensions(Inputs,Outputs)
    GenerateAssets(Inputs,Outputs)
    GenerateColorFamily(Outputs)
    GenerateSinks(Inputs,Outputs)
    #Generators

    #Save
    print('Saving')
    LowesBathroomVanitieswithTopsWorkBook.save(filename = "Output Bathroom Vanities with Tops.xlsx")
    LowesBathroomVanitieswithoutTopsWorkBook.save(filename = "Output Bathroom Vanities without Tops.xlsx")
    #LowesMakeupVanitiesWorkBook.save(filename = "Makeup Vanities.xlsx")
    print('Saved')
    #Save

if __name__ == '__main__':
    main()
