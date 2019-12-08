#Name:          Lexora EDI
#Description:   Automate EDI
#Author:        Codnr
#Version:       5
#Require:       openpyxl==2.5.14
#Notice:        Close all files used by this program before runing.

import openpyxl

def GenerateMapping():
    Mapping = []
    Data = []
    Data.append('UPC')#Name
    Data.append(['UPC'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('GTIN')#Name
    Data.append(['GTIN','globalTradeItemNumber (GTIN)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Highest Level GTIN')#Name
    Data.append(['Highest Level GTIN'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Product Name')#Name
    Data.append(['Title','Product Name'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Model Number')#Name
    Data.append(['Model Number','MFG Part # (OEM)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Vanity Color')#Name
    Data.append(['Vanity Color','Manufacturer Color/Finish'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Mirror Included')#Name
    Data.append(['Mirror','Mirror Included'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Collection')#Name
    Data.append(['Collection','Collection Name'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Product Type')#Name
    Data.append(['Product Type'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Counter')#Name
    Data.append(['Counter'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Counter Color')#Name
    Data.append(['Counter Color','Manufacturer Top Color'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('California Proposition 65 Warning Required')#Name
    Data.append(['California Proposition 65 Warning Required'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('California Prop 65 Warning Label (Digital Asset)')#Name
    Data.append(['California Prop 65 Warning Label (Digital Asset)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('CaliforniaProposition65Compliance.pdf')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Item has Restrictions?')#Name
    Data.append(['Item has Restrictions?'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Lowe's Vendor Type')#Name
    Data.append(['Lowe's Vendor Type'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('LGS/IOS Import Vendor')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Complies with Federal Lead Guidelines')#Name
    Data.append(['Complies with Federal Lead Guidelines'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Yes')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('CARB Compliant')#Name
    Data.append(['CARB Compliant'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Yes')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Solid Wood Frame')#Name
    Data.append(['Solid Wood Frame'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Yes')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Soft Close Doors')#Name
    Data.append(['Soft Close Doors'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Yes')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Integrated Electrical Outlet')#Name
    Data.append(['Integrated Electrical Outlet'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Overflow Drain Included')#Name
    Data.append(['Overflow Drain Included'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Hardware Finish')#Name
    Data.append(['Hardware Finish'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Chrome')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Manufacturer Faucet Finish')#Name
    Data.append(['Manufacturer Faucet Finish'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('N\A')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Soft Close Drawers')#Name
    Data.append(['Soft Close Drawers'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Yes')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Medicine Cabinet Included')#Name
    Data.append(['Medicine Cabinet Included'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Side Panel Material')#Name
    Data.append(['Side Panel Material'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('N/A')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Top Material')#Name
    Data.append(['Top Material'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Marble')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Item has Restrictions?')#Name
    Data.append(['Item has Restrictions?'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Style')#Name
    Data.append(['Style'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Modern')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Sink Shape')#Name
    Data.append(['Sink Shape'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Square')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Modular Design')#Name
    Data.append(['Modular Design'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Sink Material')#Name
    Data.append(['Sink Material'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Ceramic')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Edge Profile')#Name
    Data.append(['Edge Profile'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Flat')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Faucet Mount Type')#Name
    Data.append(['Faucet Mount Type'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Single hole')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Unfinished Sides')#Name
    Data.append(['Unfinished Sides'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('None')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Adjustable Shelves')#Name
    Data.append(['Adjustable Shelves'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Warranty')#Name
    Data.append(['Warranty'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('1-year limited')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Decorative Hardware Included')#Name
    Data.append(['Decorative Hardware Included'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Max Flow Rate')#Name
    Data.append(['Max Flow Rate'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('N/A')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Sink Offset')#Name
    Data.append(['Sink Offset'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('N/A')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Hardware Color Family')#Name
    Data.append(['Hardware Color Family'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Chrome')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Dove Tail Drawer Construction')#Name
    Data.append(['Dove Tail Drawer Construction'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('N/A')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Faucet Included')#Name
    Data.append(['Faucet Included'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Interior Color')#Name
    Data.append(['Interior Color'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Matches exterior')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Vanity Top Thickness in Inches')#Name
    Data.append(['Vanity Top Thickness in Inches'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('3/4-in')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Toe Kick')#Name
    Data.append(['Toe Kick'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Lowe's Exclusive')#Name
    Data.append(['Lowe's Exclusive'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Import Home Office VBU - CAN')#Name
    Data.append(['Import Home Office VBU - CAN'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('103831')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Import Home Office VBU - USA')#Name
    Data.append(['Import Home Office VBU - USA'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('103831')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Factory Name')#Name
    Data.append(['Factory Name'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Lexora')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Factory Party ID')#Name
    Data.append(['Factory Party ID'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('0')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('GTIN Wood Percentage')#Name
    Data.append(['GTIN Wood Percentage'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('100%')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('TSCA Title VI')#Name
    Data.append(['TSCA Title VI'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Product contains no composite wood')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Hazardous Indicator')#Name
    Data.append(['Hazardous Indicator'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Component Material (IOS)')#Name
    Data.append(['Component Material (IOS)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('80% rubber wood and 20% birch')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Does this item include any textile materials?')#Name
    Data.append(['Does this item include any textile materials?'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('In what country was the fabric cut?')#Name
    Data.append(['In what country was the fabric cut?'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('CHINA')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('In what country was the fabric sewn?')#Name
    Data.append(['In what country was the fabric sewn?'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('CHINA')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Is the fabric knit, woven, or non-woven?')#Name
    Data.append(['Is the fabric knit, woven, or non-woven?'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Non-Woven')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('What is the name for the mill where the fabric was formed? (i.e., knit, woven, etc.)')#Name
    Data.append(['What is the name for the mill where the fabric was formed? (i.e., knit, woven, etc.)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('N/A')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Mill Address')#Name
    Data.append(['Mill Address'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('N/A')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Does this product consist of multiple items (set)?')#Name
    Data.append(['Does this product consist of multiple items (set)?'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('No')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Origin Country')#Name
    Data.append(['Origin Country'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('CHINA')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Shipping Port (Code)')#Name
    Data.append(['Shipping Port (Code)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('USEWR')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Shipping Port Description')#Name
    Data.append(['Shipping Port Description'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('NEWARK')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Custom HTS Code (IOS)')#Name
    Data.append(['Custom HTS Code (IOS)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('9403604000')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Item Description (IOS)')#Name
    Data.append(['Item Description (IOS)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('N/A')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Import Builder Shipping Type')#Name
    Data.append(['Import Builder Shipping Type'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('NOT APPLICABLE')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Supporting Sourcing Office Location')#Name
    Data.append(['Supporting Sourcing Office Location'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Shanghai')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Ship from VBU (Ship from Factory)')#Name
    Data.append(['Ship from VBU (Ship from Factory)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('103831')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Vendor Contact Name')#Name
    Data.append(['Vendor Contact Name'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Lexora')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Vendor Email Address')#Name
    Data.append(['Vendor Email Address'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('general@lexorahome.com')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Vendor Name')#Name
    Data.append(['Vendor Name'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Lexora')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Vendor Phone Number')#Name
    Data.append(['Vendor Phone Number'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('855-453-9672')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('MFG Name')#Name
    Data.append(['MFG Name'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Lexora Home')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Manufacturer Address')#Name
    Data.append(['Manufacturer Address'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('425 Ferry Street, Newark, NJ 07105')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Lowes Merchandising Contact Name')#Name
    Data.append(['Lowes Merchandising Contact Name'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Amanda L. Wall')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Lowes Merchandising Contact Email')#Name
    Data.append(['Lowes Merchandising Contact Email'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Amanda.L.Wall@lowes.com')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Item Description {CAN_DOM}')#Name
    Data.append(['Item Description {CAN_DOM}'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('N/A')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Custom HTS Code (IOS)')#Name
    Data.append(['Custom HTS Code (IOS)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('9403604000')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Component Material')#Name
    Data.append(['Component Material'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('80% rubber wood and 20% birch.')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Country Of Origin')#Name
    Data.append(['Country Of Origin'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('CHINA')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Is Product NAFTA Eligible (Made in USA or Mexico)')#Name
    Data.append(['Is Product NAFTA Eligible (Made in USA or Mexico)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Yes')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('What is the product’s function')#Name
    Data.append(['What is the product’s function'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Bathroom Vanity')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 1')#Name
    Data.append(['Bullet point 1','Feature - Benefit Bullet 1'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('8-stage painting and finishing procs, each finish is primed and sealed for superior moisture resistance')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 2')#Name
    Data.append(['Bullet point 2','Feature - Benefit Bullet 2'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append('Solid rubber and birch wood')#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 3')#Name
    Data.append(['Bullet point 3','Feature - Benefit Bullet 3'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 4')#Name
    Data.append(['Bullet point 4','Feature - Benefit Bullet 4'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 5')#Name
    Data.append(['Bullet point 5','Feature - Benefit Bullet 5'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 6')#Name
    Data.append(['Bullet point 6','Feature - Benefit Bullet 6'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 7')#Name
    Data.append(['Bullet point 7','Feature - Benefit Bullet 7'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 8')#Name
    Data.append(['Bullet point 8','Feature - Benefit Bullet 8'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 9')#Name
    Data.append(['Bullet point 9','Feature - Benefit Bullet 9'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 10')#Name
    Data.append(['Bullet point 10'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 11')#Name
    Data.append(['Bullet point 11'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 12')#Name
    Data.append(['Bullet point 12'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 13')#Name
    Data.append(['Bullet point 13'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 14')#Name
    Data.append(['Bullet point 14'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 15')#Name
    Data.append(['Bullet point 15'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Bullet Point 16')#Name
    Data.append(['Bullet point 16'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []

    Data.append('Net Weight Unpackaged')#Name
    Data.append(['ITEM NET Weight (w/o packaging of any kind) in Pounds'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Vanity Weight Only')#Name
    Data.append(['Vanity Weight Only'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Counter Weight Only')#Name
    Data.append(['Counter Weight Only'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Mirror Weight Only')#Name
    Data.append(['Mirror Weight Only'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Number of Drawers')#Name
    Data.append(['Bullet point 11'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Number of Doors')#Name
    Data.append(['Bullet point 10'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    return Mapping

def GetMappingData(Mapping,Name):
    for Map in Mapping:
        if Map[0] == Name:
            return Map

def LoadWorkBook(FileName):
    return openpyxl.load_workbook(FileName)

def GetSpreadSheetNames(Workbook):
    return Workbook.sheetnames

def GetSpreadSheetColumnNames(Spreadsheet,Row):
    ColumnNamesList = []
    for Int in range(1,Spreadsheet.max_column+1):
        ColumnNamesList.append(Spreadsheet[GetColLetter(Int)+str(Row)].value)
    return ColumnNamesList

def GetColLetter(Col):
    return openpyxl.utils.get_column_letter(Col)

def GenerateProductSheet(Name,Type,Sheet,Row):
    return [Name,Type,Sheet,Row,GenerateMapping()]

def MapSheet(Name,Sheet,Row,Mapping):
    ColumnNamesList = GetSpreadSheetColumnNames(Sheet,Row)
    Location = 1
    for Entry in ColumnNamesList:
        for Map in Mapping:
            for Data in Map[1]:
                if Entry == Data:
                    print("Found '"+Map[0]+"' at "+str(Location)+" in '"+Name+"'")
                    Map[2] = Location
        Location += 1
    return Mapping

def Map(ProductSheet):
    for Sheet in ProductSheet:
        Sheet[4] = MapSheet(Sheet[0],Sheet[2],Sheet[3],Sheet[4])
    return ProductSheet

def CheckType(Input,Output,Location,Row):
    InputType = Input[1]
    OutputType = Output[1]
    Type = InputType
    if Type == 'Mix':
        ProductTypeColNum = GetMappingData(Input[4],'Product Type')
        CounterColNum = GetMappingData(Input[4],'Counter')
        ProductType = Input[2][GetColLetter(ProductTypeColNum[2])+str(Row+Input[3])].value
        Counter = Input[2][GetColLetter(CounterColNum[2])+str(Row+Input[3])].value
        if ProductType == 'Vanity Set':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Base Only':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity with Mirror no Counter':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity with Counter No Mirror':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Set with Side cabinet':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Set with Side cabinets':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Side Cabinet':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Make-up table':
            Type = 'Makeup Vanities'
        if Type == 'Bathroom Vanity':
            if Counter == 'Yes':
                Type = 'Bathroom Vanities with Tops'
            elif Counter == 'No':
                Type = 'Bathroom Vanities without Tops'
            else:
                print('No Counter Column Input for Bathroom Vanity')
    if Type == OutputType:
        return True
    else:
        return False

def WriteCopy(Input,Output,Location):
    #InputName = Input[0]
    InputSheet = Input[2]
    InputSheetOffset = Input[3]
    InputMapping = Input[4]
    #OutputName = Output[0]
    OutputSheet = Output[2]
    OutputSheetOffset = Output[3]
    OutputMapping = Output[4]
    for Row in range(1,InputSheet.max_row+1-InputSheetOffset):
        if CheckType(Input,Output,Location,Row):
            Wrote = False
            InputSheetValue = InputSheet[GetColLetter(InputMapping[Location][2])+str(Row+InputSheetOffset)].value
            for WriteRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                OutputSheetValue = OutputSheet[GetColLetter(OutputMapping[Location][2])+str(WriteRow)].value
                if OutputSheetValue == None and Wrote == False:
                    OutputSheet[GetColLetter(OutputMapping[Location][2])+str(WriteRow)].value = str(InputSheetValue)
                    Wrote = True
                    break
            if Wrote == False:
                OutputSheetValue = OutputSheet[GetColLetter(OutputMapping[Location][2])+str(OutputSheet.max_row)].value
                if OutputSheetValue != None:
                    OutputSheetValue = OutputSheet[GetColLetter(OutputMapping[Location][2])+str(OutputSheet.max_row+1)].value
                    OutputSheet[GetColLetter(OutputMapping[Location][2])+str(OutputSheet.max_row)].value = str(InputSheetValue)
            #print(InputName,InputMapping[Location][0],InputSheetValue,OutputName,OutputMapping[Location][0],OutputSheetValue)

def AutoFill(Inputs,Outputs):
    for Input in Inputs:
        for Output in Outputs:
            if Input[1] == 'Mix' or Input[1] == Output[1]:
                Location = 0
                for Map in Input[4]:
                    if Map[2] != None and Output[4][Location][2] != None and Map[3] == True:
                        print("Copy '"+Input[0]+"' "+Map[0]+" ["+GetColLetter(Map[2])+"] to '"+Output[0]+"' "+Output[4][Location][0]+" ["+GetColLetter(Output[4][Location][2])+"]")
                        WriteCopy(Input,Output,Location)
                    Location += 1

def WriteDefault(Output,Location):
    Name = Output[0]
    Sheet = Output[2]
    SheetOffset = Output[3]
    Mapping = Output[4]
    for Row in range(1,Sheet.max_row+1-SheetOffset):
        ExistingData = Sheet[GetColLetter(Mapping[Location][2])+str(Row+SheetOffset)].value
        WrittingData = Mapping[Location][4]
        if WrittingData != None:
            if ExistingData != None:
                Sheet[GetColLetter(Mapping[Location][2])+str(Row+SheetOffset)].value = WrittingData
                #print(Name+" "+Mapping[Location][0]+" Row: "+str(Row+SheetOffset)+" Overwrote Existing Data '"+str(ExistingData)+"' WrittingData '"+str(WrittingData)+"'")
            else:
                Sheet[GetColLetter(Mapping[Location][2])+str(Row+SheetOffset)].value = WrittingData
                #print(Name+" "+Mapping[Location][0]+" Row: "+str(Row+SheetOffset)+" Wrote '"+str(WrittingData)+"'")
        else:
            print(Name+" "+" Row: "+str(Row+SheetOffset)+" WrittingData is 'None'")

def AutoDefault(Outputs):
    for Output in Outputs:
        Location = 0
        for Map in Output[4]:
            if Map[2] != None and Map[4] != None:
                print("Defaulting '"+Output[0]+"' "+Map[0]+" ["+GetColLetter(Map[2])+"] to '"+str(Map[4])+"'")
                WriteDefault(Output,Location)
            Location += 1

def CalculateNetWeightUnpackaged(VanityWeightOnly,CounterWeightOnly,MirrorWeightOnly):
    return str(round(float(str(VanityWeightOnly).replace('N/A','0'))+float(str(CounterWeightOnly).replace('N/A','0'))+float(str(MirrorWeightOnly).replace('N/A','0')),2))

def GenerateNetWeightUnpackaged(Inputs,Outputs):
    for Input in Inputs:
        InputName = Input[0]
        InputSheet = Input[2]
        InputSheetOffset = Input[3]
        InputMapping = Input[4]
        for Output in Outputs:
            OutputName = Output[0]
            OutputSheet = Output[2]
            OutputSheetOffset = Output[3]
            OutputMapping = Output[4]
            NetWeightUnpackaged = GetMappingData(OutputMapping,'Net Weight Unpackaged')
            OutputUPC = GetMappingData(OutputMapping,'UPC')
            InputUPC = GetMappingData(InputMapping,'UPC')
            VanityWeightOnly = GetMappingData(InputMapping,'Vanity Weight Only')
            CounterWeightOnly = GetMappingData(InputMapping,'Counter Weight Only')
            MirrorWeightOnly = GetMappingData(InputMapping,'Mirror Weight Only')
            if NetWeightUnpackaged != None and OutputUPC != None and InputUPC != None and VanityWeightOnly != None and CounterWeightOnly != None and MirrorWeightOnly != None:
                Msg = "Generating '"+OutputName+"' "+NetWeightUnpackaged[0]+" ["+GetColLetter(NetWeightUnpackaged[2])+"] using '"+InputName+"' "+VanityWeightOnly[0]+" ["+GetColLetter(VanityWeightOnly[2])+"] & "
                Msg += CounterWeightOnly[0]+" ["+GetColLetter(CounterWeightOnly[2])+"] & "+MirrorWeightOnly[0]+" ["+GetColLetter(MirrorWeightOnly[2])+"]"
                print(Msg)
                for OutputRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                    OutputUPCValue = OutputSheet[GetColLetter(OutputUPC[2])+str(OutputRow)].value
                    Wrote = False
                    for InputRow in range(1+InputSheetOffset,InputSheet.max_row+1):
                        InputUPCValue = str(InputSheet[GetColLetter(InputUPC[2])+str(InputRow)].value)
                        if OutputUPCValue != None and InputUPCValue != None and str(OutputUPCValue) == str(InputUPCValue):
                            OutputNetWeightUnpackaged = OutputSheet[GetColLetter(NetWeightUnpackaged[2])+str(OutputRow)].value
                            InputVanityWeightOnly = InputSheet[GetColLetter(VanityWeightOnly[2])+str(InputRow)].value
                            InputCounterWeightOnly = InputSheet[GetColLetter(CounterWeightOnly[2])+str(InputRow)].value
                            InputMirrorWeightOnly = InputSheet[GetColLetter(MirrorWeightOnly[2])+str(InputRow)].value
                            if InputVanityWeightOnly != None and InputCounterWeightOnly != None and InputMirrorWeightOnly != None:
                                CalculatedNetWeightUnpackaged = CalculateNetWeightUnpackaged(InputVanityWeightOnly,InputCounterWeightOnly,InputMirrorWeightOnly)
                            else:
                                print("Error '"+OutputName+"' Row: "+str(OutputRow)+" Can't Calculate "+NetWeightUnpackaged[0]+" ["+GetColLetter(NetWeightUnpackaged[2])+"] "+"of '"+InputName+"' Row: "+str(InputRow))
                                CalculatedNetWeightUnpackaged = 'Missing'
                            if OutputNetWeightUnpackaged != None:
                                OutputSheet[GetColLetter(NetWeightUnpackaged[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Net Weight Unpackaged of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(OutputNetWeightUnpackaged)+" New: "+CalculatedNetWeightUnpackaged)
                                Wrote = True
                                break
                            else:
                                OutputSheet[GetColLetter(NetWeightUnpackaged[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                #print(OutputName+" Row: "+str(OutputRow)+" Wrote Net Weight Unpackaged of UPC: "+str(OutputUPCValue)+" to "+CalculatedNetWeightUnpackaged)
                                Wrote = True
                                break
                        else:
                            if OutputUPCValue == None:
                                print("Error '"+OutputName+"' Row: "+str(OutputRow)+" No UPC")
                                break
                    if Wrote == False:
                        print("Error '"+OutputName+"' Unable to find UPC: "+str(OutputSheet[GetColLetter(OutputUPC[2])+str(OutputRow)].value))

def GenerateGTIN(Outputs):
    for Output in Outputs:
        OutputName = Output[0]
        OutputSheet = Output[2]
        OutputSheetOffset = Output[3]
        OutputMapping = Output[4]
        UPC = GetMappingData(OutputMapping,'UPC')
        GTIN = GetMappingData(OutputMapping,'GTIN')
        HighestLevelGTIN = GetMappingData(OutputMapping,'Highest Level GTIN')
        Location = 0
        if UPC != None and UPC[2] != None and GTIN and GTIN[2] != None and HighestLevelGTIN and HighestLevelGTIN[2] != None:
            OutputUPC = OutputSheet[GetColLetter(UPC[2])+str(InputRow)].value
            OutputGTIN = OutputSheet[GetColLetter(GTIN[2])+str(InputRow)].value
            OutputHighestLevelGTIN = OutputSheet[GetColLetter(HighestLevelGTIN[2])+str(InputRow)].value
        for Map in Output[4]:
            if Map[2] != None and Map[4] != None:
                print("Defaulting '"+Output[0]+"' "+Map[0]+" ["+GetColLetter(Map[2])+"] to '"+str(Map[4])+"'")
                WriteDefault(Output,Location)
            Location += 1
    return '00'+UPC

def main():
    #Load Input Workbooks
    LexoraSourceWorkBook = LoadWorkBook("Lexora Source.xlsx")
    #Load Input Workbooks

    #Load Output Workbooks
    LowesBathroomVanitieswithTopsWorkBook = LoadWorkBook("Bathroom Vanities with Tops.xlsx")
    LowesBathroomVanitieswithoutTopsWorkBook = LoadWorkBook("Bathroom Vanities without Tops.xlsx")
    ###LowesMakeupVanitiesWorkBook = LoadWorkBook("Makeup Vanities.xlsx")
    #Load Output Workbooks

    #Load Input Sheets
    BathroomVanities = LexoraSourceWorkBook['Bathroom Vanities']
    #Load Input Sheets

    #Load Output Sheets
    BathroomVanitieswithTops = LowesBathroomVanitieswithTopsWorkBook['Attributes']
    BathroomVanitieswithoutTops = LowesBathroomVanitieswithoutTopsWorkBook['Attributes']
    ###MakeupVanities = LowesMakeupVanitiesWorkBook['Attributes']
    #Load Output Sheets

    #Load Input List
    Inputs = []
    Inputs.append(GenerateProductSheet('Bathroom Vanities','Mix',BathroomVanities,1))
    #Load Input List

    #Load Output List
    Outputs = []
    Outputs.append(GenerateProductSheet('Bathroom Vanities with Tops','Bathroom Vanities with Tops',BathroomVanitieswithTops,5))
    Outputs.append(GenerateProductSheet('Bathroom Vanities without Tops','Bathroom Vanities without Tops',BathroomVanitieswithoutTops,5))
    ###Outputs.append(GenerateProductSheet('Makeup Vanities','Makeup Vanities',MakeupVanities,5))
    #Load Output List

    #Mapping
    Inputs = Map(Inputs)
    Outputs = Map(Outputs)
    #Mapping

    AutoFill(Inputs,Outputs)
    AutoDefault(Outputs)
    GenerateNetWeightUnpackaged(Inputs,Outputs)
    #GenerateGTIN(Outputs)

    #Save
    LowesBathroomVanitieswithTopsWorkBook.save(filename = "Output Bathroom Vanities with Tops.xlsx")
    LowesBathroomVanitieswithoutTopsWorkBook.save(filename = "Output Bathroom Vanities without Tops.xlsx")
    #LowesMakeupVanitiesWorkBook.save(filename = "Makeup Vanities.xlsx")
    #Save

if __name__ == '__main__':
    main()