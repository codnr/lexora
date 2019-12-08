#Name:          Lexora EDI
#Description:   Automate EDI
#Author:        Codnr
#Version:       2
#Require:       openpyxl==2.5.14
#Notice:        Close all files used by this program before runing.

import openpyxl

def GenerateMapping():
    Mapping = []
    Data = []
    Data.append('UPC')#Name
    Data.append(['UPC'])#Alternate name list
    Data.append(None)#Recieve Column
    Data.append(None)#Transfer Column
    Mapping.append(Data)
    Data = []
    return Mapping

def GenerateGTIN(UPC):
    return '00'+UPC

def GetSpreadSheetColumnNames(Spreadsheet,Row):
    ColumnNamesList = []
    for Int in range(1,Spreadsheet.max_column+1):
        ColumnNamesList.append(Spreadsheet[openpyxl.utils.get_column_letter(Int)+'1'].value)
    return ColumnNamesList

def main():
    #
    Mapping = GenerateMapping()
    #

    #Read
    LexoraSpreadSheet = openpyxl.load_workbook("LEXORA Master DATA FILE_FOR_DEALERS_6-25-19.xlsx")
    LexoraSpreadSheetNames = ['Fireplaces']#LexoraSpreadSheet.sheetnames
    print(LexoraSpreadSheetNames)
    #BathroomVanities = LexoraSpreadSheet['Bathroom Vanities']
    #BathtubsandFaucets = LexoraSpreadSheet['Bathtubs&Faucets']
    #Fireplaces = LexoraSpreadSheet['Fireplaces']
    #UpdateHistory = LexoraSpreadSheet['Update History']
    for SheetNames in LexoraSpreadSheetNames:
        ActiveSheet = LexoraSpreadSheet[SheetNames]
        print('Sheet: '+SheetNames+' - Max Columns: '+str(ActiveSheet.max_column)+'/'+openpyxl.utils.get_column_letter(ActiveSheet.max_column))
        ColumnNamesList = GetSpreadSheetColumnNames(ActiveSheet,'1')
        Location = 1
        for Entry in ColumnNamesList:
            for Map in Mapping:
                for Data in Map[1]:
                    if Entry == Data:
                        print('Found '+Map[0]+' at '+str(Location))
                        Map[2] = Location
            Location += 1
    
    return

    #Read


    #Write
    LexoraLowesSpreadSheet = openpyxl.load_workbook("LexoraHome-Lowe's-Lowe's-CoreMarketing-Lowe's-BathroomVanitieswith-08082019.xlsx")
    #Summary = LexoraLowesSpreadSheet['Summary']
    Attributes = LexoraLowesSpreadSheet['Attributes']
    #Reference = LexoraLowesSpreadSheet['Reference']
    WUPC = Attributes['DW6'].value
    if WUPC == None:
        Attributes['DW6'] = RUPC
        print('Written')
        print('WUPC: ',Attributes['DW6'].value)
    else:
        print('Existing Value')
        print('WUPC: ',WUPC)
    LexoraLowesSpreadSheet.save(filename = "LexoraHome-Lowe's-Lowe's-CoreMarketing-Lowe's-BathroomVanitieswith-08082019.xlsx")
    #Write


if __name__ == '__main__':
    main()