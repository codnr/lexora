#Name:          Lexora EDI
#Description:   Automate EDI
#Author:        Codnr
#Version:       3
#Require:       openpyxl==2.5.14
#Notice:        Close all files used by this program before runing.

import openpyxl

def GenerateMapping():
    Mapping = []
    Data = []
    Data.append('Product Type')#Name
    Data.append(['Product Type'])#Alternate name list
    Data.append(None)#Recieve Column
    Data.append(None)#Transfer Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('UPC')#Name
    Data.append(['UPC'])#Alternate name list
    Data.append(None)#Recieve Column
    Data.append(None)#Transfer Column
    Data.append(True)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('GTIN')#Name
    Data.append(['GTIN','globalTradeItemNumber (GTIN)'])#Alternate name list
    Data.append(None)#Recieve Column
    Data.append(None)#Transfer Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Net Weight Unpackaged')#Name
    Data.append(['ITEM NET Weight (w/o packaging of any kind) in Pounds'])#Alternate name list
    Data.append(None)#Recieve Column
    Data.append(None)#Transfer Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Vanity Weight Only')#Name
    Data.append(['Vanity Weight Only'])#Alternate name list
    Data.append(None)#Recieve Column
    Data.append(None)#Transfer Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Counter Weight Only')#Name
    Data.append(['Counter Weight Only'])#Alternate name list
    Data.append(None)#Recieve Column
    Data.append(None)#Transfer Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Mirror Weight Only')#Name
    Data.append(['Mirror Weight Only'])#Alternate name list
    Data.append(None)#Recieve Column
    Data.append(None)#Transfer Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Number of Drawers')#Name
    Data.append(['Bullet point 11'])#Alternate name list
    Data.append(None)#Recieve Column
    Data.append(None)#Transfer Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Number of Doors')#Name
    Data.append(['Bullet point 10'])#Alternate name list
    Data.append(None)#Recieve Column
    Data.append(None)#Transfer Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Counter Color')#Name
    Data.append(['Counter Color','Manufacturer Top Color'])#Alternate name list
    Data.append(None)#Recieve Column
    Data.append(None)#Transfer Column
    Data.append(True)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    return Mapping

def GetMappingData(Mapping,Name):
    for Map in Mapping:
        if Map[0] == Name:
            return Map

def LoadWorkBook(FileName):
    return openpyxl.load_workbook(FileName)

def GetSpreadSheetNames(Workbook):
    return Workbook.sheetnames

def GetSpreadSheetColumnNames(Spreadsheet,Row):
    ColumnNamesList = []
    for Int in range(1,Spreadsheet.max_column+1):
        ColumnNamesList.append(Spreadsheet[openpyxl.utils.get_column_letter(Int)+str(Row)].value)
    return ColumnNamesList

def MapOutputSheet(Mapping,OutputSheet):
    ColumnNamesList = GetSpreadSheetColumnNames(OutputSheet,5)
    Location = 1
    for Entry in ColumnNamesList:
        for Map in Mapping:
            for Data in Map[1]:
                if Entry == Data:
                    print("Found Output '"+Map[0]+"' at "+str(Location))
                    Map[3] = Location
        Location += 1
    return Mapping

def MapInputSheet(Mapping,InputSheet):
    ColumnNamesList = GetSpreadSheetColumnNames(InputSheet,1)
    Location = 1
    for Entry in ColumnNamesList:
        for Map in Mapping:
            for Data in Map[1]:
                if Entry == Data:
                    print("Found Input '"+Map[0]+"' at "+str(Location))
                    Map[2] = Location
        Location += 1
    return Mapping

def WriteCopy(InputSheet,OutputSheet,Map):
    InputSheetOffset = 1
    OutputSheetOffset = 5
    for Row in range(1,OutputSheet.max_row+1-OutputSheetOffset):
        Input = InputSheet[openpyxl.utils.get_column_letter(Map[2])+str(Row+InputSheetOffset)].value
        Output = OutputSheet[openpyxl.utils.get_column_letter(Map[3])+str(Row+OutputSheetOffset)].value
        print(Input,Output)
        continue
        if OutputUPC != None and OutputUPC == InputUPC:
            if OutputNetWeightUnpackaged != None:
                OutputSheet[openpyxl.utils.get_column_letter(NetWeightUnpackaged[3])+str(Row+OutputSheetOffset)].value = CalculatedNetWeightUnpackaged
                print("Row: "+str(Row+OutputSheetOffset)+" Overwrote Net Weight Unpackaged of UPC: "+str(OutputUPC)+" to "+"Old: "+str(OutputNetWeightUnpackaged)+" New: "+CalculatedNetWeightUnpackaged)
            else:
                OutputSheet[openpyxl.utils.get_column_letter(NetWeightUnpackaged[3])+str(Row+OutputSheetOffset)].value = CalculatedNetWeightUnpackaged
                print("Row: "+str(Row+OutputSheetOffset)+" Wrote Net Weight Unpackaged of UPC: "+str(OutputUPC)+" to "+CalculatedNetWeightUnpackaged)
        else:
            if OutputUPC == None:
                print("Row: "+str(Row+OutputSheetOffset)+" No UPC")
            elif OutputUPC != InputUPC:
                print("Row: "+str(Row+OutputSheetOffset)+" UPC Mismatch Input UPC: "+str(InputUPC)+" Output UPC: "+str(OutputUPC))
            else:
                print("Row: "+str(Row+OutputSheetOffset)+" Unspecified Error Check Sheets")

def AutoFill(Mapping,InputSheet,OutputSheet):
    for Map in Mapping:
        if Map[3] != None and Map[4] != None:
            pass#Write default
        elif Map[2] != None and Map[3] != None:
            WriteCopy(InputSheet,OutputSheet,Map)
    #Save

#def GenerateGTIN(UPC):
#    return '00'+UPC

def CalculateNetWeightUnpackaged(VanityWeightOnly,CounterWeightOnly,MirrorWeightOnly):
    return str(round(float(str(VanityWeightOnly).replace('N/A','0'))+float(str(CounterWeightOnly).replace('N/A','0'))+float(str(MirrorWeightOnly).replace('N/A','0')),2))


def GenerateNetWeightUnpackaged(Mapping,InputSheet,OutputSheet):
    UPC = GetMappingData(Mapping,'UPC')
    NetWeightUnpackaged = GetMappingData(Mapping,'Net Weight Unpackaged')
    VanityWeightOnly = GetMappingData(Mapping,'Vanity Weight Only')
    CounterWeightOnly = GetMappingData(Mapping,'Counter Weight Only')
    MirrorWeightOnly = GetMappingData(Mapping,'Mirror Weight Only')
    InputSheetOffset = 1
    OutputSheetOffset = 5
    for Row in range(1,OutputSheet.max_row+1-OutputSheetOffset):
        InputUPC = InputSheet[openpyxl.utils.get_column_letter(UPC[2])+str(Row+InputSheetOffset)].value
        OutputUPC = OutputSheet[openpyxl.utils.get_column_letter(UPC[3])+str(Row+OutputSheetOffset)].value
        InputVanityWeightOnly = InputSheet[openpyxl.utils.get_column_letter(VanityWeightOnly[2])+str(Row+InputSheetOffset)].value
        InputCounterWeightOnly = InputSheet[openpyxl.utils.get_column_letter(CounterWeightOnly[2])+str(Row+InputSheetOffset)].value
        InputMirrorWeightOnly = InputSheet[openpyxl.utils.get_column_letter(MirrorWeightOnly[2])+str(Row+InputSheetOffset)].value
        OutputNetWeightUnpackaged = OutputSheet[openpyxl.utils.get_column_letter(NetWeightUnpackaged[3])+str(Row+OutputSheetOffset)].value
        CalculatedNetWeightUnpackaged = CalculateNetWeightUnpackaged(InputVanityWeightOnly,InputCounterWeightOnly,InputMirrorWeightOnly)
        if OutputUPC != None and OutputUPC == InputUPC:
            if OutputNetWeightUnpackaged != None:
                OutputSheet[openpyxl.utils.get_column_letter(NetWeightUnpackaged[3])+str(Row+OutputSheetOffset)].value = CalculatedNetWeightUnpackaged
                print("Row: "+str(Row+OutputSheetOffset)+" Overwrote Net Weight Unpackaged of UPC: "+str(OutputUPC)+" to "+"Old: "+str(OutputNetWeightUnpackaged)+" New: "+CalculatedNetWeightUnpackaged)
            else:
                OutputSheet[openpyxl.utils.get_column_letter(NetWeightUnpackaged[3])+str(Row+OutputSheetOffset)].value = CalculatedNetWeightUnpackaged
                print("Row: "+str(Row+OutputSheetOffset)+" Wrote Net Weight Unpackaged of UPC: "+str(OutputUPC)+" to "+CalculatedNetWeightUnpackaged)
        else:
            if OutputUPC == None:
                print("Row: "+str(Row+OutputSheetOffset)+" No UPC")
            elif OutputUPC != InputUPC:
                print("Row: "+str(Row+OutputSheetOffset)+" UPC Mismatch Input UPC: "+str(InputUPC)+" Output UPC: "+str(OutputUPC))
            else:
                print("Row: "+str(Row+OutputSheetOffset)+" Unspecified Error Check Sheets")

def main():
    Mapping = GenerateMapping()
    LexoraWorkBook = LoadWorkBook("LEXORA Master DATA FILE_FOR_DEALERS_6-25-19.xlsx")
    LexoraLowesWorkBook = LoadWorkBook("test.xlsx")
    InputSheet = LexoraWorkBook['Bathroom Vanities']
    OutputSheet = LexoraLowesWorkBook['Attributes']
    Mapping = MapInputSheet(Mapping,InputSheet)
    Mapping = MapOutputSheet(Mapping,OutputSheet)
    #AutoFill(Mapping,InputSheet,OutputSheet)
    GenerateNetWeightUnpackaged(Mapping,InputSheet,OutputSheet)
    LexoraLowesWorkBook.save(filename = "test.xlsx")
    #print(InputSheet.max_column)
    #print(InputSheet.max_row)
    #print(OutputSheet.max_column)
    #print(OutputSheet.max_row)

    return

    #Read


    #Write
    LexoraLowesSpreadSheet = openpyxl.load_workbook("LexoraHome-Lowe's-Lowe'sInitiateUSA-08082019.xlsx")
    #Summary = LexoraLowesSpreadSheet['Summary']
    Attributes = LexoraLowesSpreadSheet['Attributes']
    #Reference = LexoraLowesSpreadSheet['Reference']
    WUPC = Attributes['DW6'].value
    if WUPC == None:
        Attributes['DW6'] = RUPC
        print('Written')
        print('WUPC: ',Attributes['DW6'].value)
    else:
        print('Existing Value')
        print('WUPC: ',WUPC)
    LexoraLowesSpreadSheet.save(filename = "LexoraHome-Lowe's-Lowe'sInitiateUSA-08082019.xlsx")
    #Write


if __name__ == '__main__':
    main()