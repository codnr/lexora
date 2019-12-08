#Name:          Lexora EDI
#Description:   Automate EDI
#Author:        Codnr
#Version:       4
#Require:       openpyxl==2.5.14
#Notice:        Close all files used by this program before runing.

import openpyxl

def GenerateMapping():
    Mapping = []
    Data = []
    Data.append('UPC')#Name
    Data.append(['UPC'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('GTIN')#Name
    Data.append(['GTIN','globalTradeItemNumber (GTIN)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Product Name')#Name
    Data.append(['Title','Product Name'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Model Number')#Name
    Data.append(['Model Number','MFG Part # (OEM)'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Vanity Color')#Name
    Data.append(['Vanity Color','Manufacturer Color/Finish'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Mirror Included')#Name
    Data.append(['Mirror','Mirror Included'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Collection')#Name
    Data.append(['Collection','Collection Name'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Product Type')#Name
    Data.append(['Product Type'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Counter')#Name
    Data.append(['Counter'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Counter Color')#Name
    Data.append(['Counter Color','Manufacturer Top Color'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Net Weight Unpackaged')#Name
    Data.append(['ITEM NET Weight (w/o packaging of any kind) in Pounds'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Vanity Weight Only')#Name
    Data.append(['Vanity Weight Only'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Counter Weight Only')#Name
    Data.append(['Counter Weight Only'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Mirror Weight Only')#Name
    Data.append(['Mirror Weight Only'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Number of Drawers')#Name
    Data.append(['Bullet point 11'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append('Number of Doors')#Name
    Data.append(['Bullet point 10'])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    return Mapping

def GetMappingData(Mapping,Name):
    for Map in Mapping:
        if Map[0] == Name:
            return Map

def LoadWorkBook(FileName):
    return openpyxl.load_workbook(FileName)

def GetSpreadSheetNames(Workbook):
    return Workbook.sheetnames

def GetSpreadSheetColumnNames(Spreadsheet,Row):
    ColumnNamesList = []
    for Int in range(1,Spreadsheet.max_column+1):
        ColumnNamesList.append(Spreadsheet[openpyxl.utils.get_column_letter(Int)+str(Row)].value)
    return ColumnNamesList

def GenerateProductSheet(Name,Type,Sheet,Row):
    return [Name,Type,Sheet,Row,GenerateMapping()]

def MapSheet(Name,Sheet,Row,Mapping):
    ColumnNamesList = GetSpreadSheetColumnNames(Sheet,Row)
    Location = 1
    for Entry in ColumnNamesList:
        for Map in Mapping:
            for Data in Map[1]:
                if Entry == Data:
                    print("Found '"+Map[0]+"' at "+str(Location)+" in '"+Name+"'")
                    Map[2] = Location
        Location += 1
    return Mapping

def Map(ProductSheet):
    for Sheet in ProductSheet:
        Sheet[4] = MapSheet(Sheet[0],Sheet[2],Sheet[3],Sheet[4])
    return ProductSheet

def CheckType(Input,Output,Location,Row):
    InputType = Input[1]
    OutputType = Output[1]
    Type = InputType
    if Type == 'Mix':
        ProductTypeColNum = GetMappingData(Input[4],'Product Type')
        CounterColNum = GetMappingData(Input[4],'Counter')
        ProductType = Input[2][openpyxl.utils.get_column_letter(ProductTypeColNum[2])+str(Row+Input[3])].value
        Counter = Input[2][openpyxl.utils.get_column_letter(CounterColNum[2])+str(Row+Input[3])].value
        if ProductType == 'Vanity Set':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Base Only':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity with Mirror no Counter':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity with Counter No Mirror':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Set with Side cabinet':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Set with Side cabinets':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Side Cabinet':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Make-up table':
            Type = 'Makeup Vanities'
        if Type == 'Bathroom Vanity':
            if Counter == 'Yes':
                Type = 'Bathroom Vanities with Tops'
            elif Counter == 'No':
                Type = 'Bathroom Vanities without Tops'
            else:
                print('No Counter Column Input for Bathroom Vanity')
    if Type == OutputType:
        return True
    else:
        return False

def WriteDefault(Output,Location):
    Name = Output[0]
    Sheet = Output[2]
    SheetOffset = Output[3]
    Mapping = Output[4]
    for Row in range(1,Sheet.max_row+1-SheetOffset):
        ExistingData = Sheet[openpyxl.utils.get_column_letter(Location)+str(Row+SheetOffset)].value
        WrittingData = Mapping[Location][5]
        print(Name+" "+Mapping[Location][0]+" Row: "+str(Row+SheetOffset)+" Existing Data '"+ExistingData+"' WrittingData "+WrittingData)
        continue
        if OutputUPC != None and OutputUPC == InputUPC:
            if OutputNetWeightUnpackaged != None:
                OutputSheet[openpyxl.utils.get_column_letter(NetWeightUnpackaged[3])+str(Row+OutputSheetOffset)].value = CalculatedNetWeightUnpackaged
                print("Row: "+str(Row+OutputSheetOffset)+" Overwrote Net Weight Unpackaged of UPC: "+str(OutputUPC)+" to "+"Old: "+str(OutputNetWeightUnpackaged)+" New: "+CalculatedNetWeightUnpackaged)
            else:
                OutputSheet[openpyxl.utils.get_column_letter(NetWeightUnpackaged[3])+str(Row+OutputSheetOffset)].value = CalculatedNetWeightUnpackaged
                print("Row: "+str(Row+OutputSheetOffset)+" Wrote Net Weight Unpackaged of UPC: "+str(OutputUPC)+" to "+CalculatedNetWeightUnpackaged)
        else:
            if OutputUPC == None:
                print("Row: "+str(Row+OutputSheetOffset)+" No UPC")
            elif OutputUPC != InputUPC:
                print("Row: "+str(Row+OutputSheetOffset)+" UPC Mismatch Input UPC: "+str(InputUPC)+" Output UPC: "+str(OutputUPC))
            else:
                print("Row: "+str(Row+OutputSheetOffset)+" Unspecified Error Check Sheets")

def WriteCopy(Input,Output,Location):
    InputName = Input[0]
    InputSheet = Input[2]
    InputSheetOffset = Input[3]
    InputMapping = Input[4]
    OutputName = Output[0]
    OutputSheet = Output[2]
    OutputSheetOffset = Output[3]
    OutputMapping = Output[4]
    for Row in range(1,InputSheet.max_row+1-InputSheetOffset):
        if CheckType(Input,Output,Location,Row):
            Wrote = False
            InputSheetValue = InputSheet[openpyxl.utils.get_column_letter(InputMapping[Location][2])+str(Row+InputSheetOffset)].value
            for WriteRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                OutputSheetValue = OutputSheet[openpyxl.utils.get_column_letter(OutputMapping[Location][2])+str(WriteRow)].value
                if OutputSheetValue == None and Wrote == False:
                    OutputSheet[openpyxl.utils.get_column_letter(OutputMapping[Location][2])+str(WriteRow)].value = str(InputSheetValue)
                    Wrote = True
                    break
            if Wrote == False:
                OutputSheetValue = OutputSheet[openpyxl.utils.get_column_letter(OutputMapping[Location][2])+str(OutputSheet.max_row)].value
                if OutputSheetValue != None:
                    OutputSheetValue = OutputSheet[openpyxl.utils.get_column_letter(OutputMapping[Location][2])+str(OutputSheet.max_row+1)].value
                    OutputSheet[openpyxl.utils.get_column_letter(OutputMapping[Location][2])+str(OutputSheet.max_row)].value = str(InputSheetValue)
            #print(InputName,InputMapping[Location][0],InputSheetValue,OutputName,OutputMapping[Location][0],OutputSheetValue)

def AutoFill(Inputs,Outputs):
    for Input in Inputs:
        for Output in Outputs:
            if Input[1] == 'Mix' or Input[1] == Output[1]:
                Location = 0
                for Map in Input[4]:
                    if Map[2] != None and Output[4][Location][2] != None and Map[3] == True:
                        print("Copy "+Input[0]+" "+Map[0]+" ["+openpyxl.utils.get_column_letter(Map[2])+"] to "+Output[0]+" "+Output[4][Location][0]+" ["+openpyxl.utils.get_column_letter(Output[4][Location][2])+"]")
                        WriteCopy(Input,Output,Location)
                    Location += 1

def AutoDefault(Outputs):
    pass
#if Output[4][Location][2] != None and Output[4][Location][4] != None:
#    WriteDefault(Output,Location)


#def GenerateGTIN(UPC):
#    return '00'+UPC

def CalculateNetWeightUnpackaged(VanityWeightOnly,CounterWeightOnly,MirrorWeightOnly):
    return str(round(float(str(VanityWeightOnly).replace('N/A','0'))+float(str(CounterWeightOnly).replace('N/A','0'))+float(str(MirrorWeightOnly).replace('N/A','0')),2))


def GenerateNetWeightUnpackaged(Inputs,Outputs):
    for Input in Inputs:
        InputName = Input[0]
        InputSheet = Input[2]
        InputSheetOffset = Input[3]
        InputMapping = Input[4]
        for Output in Outputs:
            OutputName = Output[0]
            OutputSheet = Output[2]
            OutputSheetOffset = Output[3]
            OutputMapping = Output[4]
            NetWeightUnpackaged = GetMappingData(OutputMapping,'Net Weight Unpackaged')
            OutputUPC = GetMappingData(OutputMapping,'UPC')
            InputUPC = GetMappingData(InputMapping,'UPC')
            VanityWeightOnly = GetMappingData(InputMapping,'Vanity Weight Only')
            CounterWeightOnly = GetMappingData(InputMapping,'Counter Weight Only')
            MirrorWeightOnly = GetMappingData(InputMapping,'Mirror Weight Only')
            if NetWeightUnpackaged != None and OutputUPC != None and InputUPC != None and VanityWeightOnly != None and CounterWeightOnly != None and MirrorWeightOnly != None:
                for OutputRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                    OutputUPCValue = OutputSheet[openpyxl.utils.get_column_letter(OutputUPC[2])+str(OutputRow)].value
                    Wrote = False
                    for InputRow in range(1+InputSheetOffset,InputSheet.max_row+1):
                        InputUPCValue = str(InputSheet[openpyxl.utils.get_column_letter(InputUPC[2])+str(InputRow)].value)
                        if OutputUPCValue != None and InputUPCValue != None and str(OutputUPCValue) == str(InputUPCValue):
                            OutputNetWeightUnpackaged = OutputSheet[openpyxl.utils.get_column_letter(NetWeightUnpackaged[2])+str(OutputRow)].value
                            InputVanityWeightOnly = InputSheet[openpyxl.utils.get_column_letter(VanityWeightOnly[2])+str(InputRow)].value
                            InputCounterWeightOnly = InputSheet[openpyxl.utils.get_column_letter(CounterWeightOnly[2])+str(InputRow)].value
                            InputMirrorWeightOnly = InputSheet[openpyxl.utils.get_column_letter(MirrorWeightOnly[2])+str(InputRow)].value
                            if InputVanityWeightOnly != None and InputCounterWeightOnly != None and InputMirrorWeightOnly != None:
                                CalculatedNetWeightUnpackaged = CalculateNetWeightUnpackaged(InputVanityWeightOnly,InputCounterWeightOnly,InputMirrorWeightOnly)
                            else:
                                print(OutputName+" Row: "+str(OutputRow)+" Can't Calculate Net Weight Unpackaged of "+OutputName+" Row: "+str(InputRow))
                                CalculatedNetWeightUnpackaged = 'Missing'
                            if OutputNetWeightUnpackaged != None:
                                OutputSheet[openpyxl.utils.get_column_letter(NetWeightUnpackaged[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                print(OutputName+" Row: "+str(OutputRow)+" Overwrote Net Weight Unpackaged of UPC: "+str(OutputUPCValue)+" to "+"Old: "+str(OutputNetWeightUnpackaged)+" New: "+CalculatedNetWeightUnpackaged)
                                Wrote = True
                                break
                            else:
                                OutputSheet[openpyxl.utils.get_column_letter(NetWeightUnpackaged[2])+str(OutputRow)].value = CalculatedNetWeightUnpackaged
                                print(OutputName+" Row: "+str(OutputRow)+" Wrote Net Weight Unpackaged of UPC: "+str(OutputUPCValue)+" to "+CalculatedNetWeightUnpackaged)
                                Wrote = True
                                break
                        else:
                            if OutputUPCValue == None:
                                print(OutputName+" Row: "+str(OutputRow)+" No UPC")
                                break
                    if Wrote == False:
                        print('Unable to find UPC: '+str(OutputSheet[openpyxl.utils.get_column_letter(OutputUPC[2])+str(OutputRow)].value))

def main():
    #Load Input Workbooks
    LexoraSourceWorkBook = LoadWorkBook("Lexora Source.xlsx")
    #Load Input Workbooks

    #Load Output Workbooks
    LowesBathroomVanitieswithTopsWorkBook = LoadWorkBook("Bathroom Vanities with Tops.xlsx")
    LowesBathroomVanitieswithoutTopsWorkBook = LoadWorkBook("Bathroom Vanities without Tops.xlsx")
    ###LowesMakeupVanitiesWorkBook = LoadWorkBook("Makeup Vanities.xlsx")
    #Load Output Workbooks

    #Load Input Sheets
    BathroomVanities = LexoraSourceWorkBook['Bathroom Vanities']
    #Load Input Sheets

    #Load Output Sheets
    BathroomVanitieswithTops = LowesBathroomVanitieswithTopsWorkBook['Attributes']
    BathroomVanitieswithoutTops = LowesBathroomVanitieswithoutTopsWorkBook['Attributes']
    ###MakeupVanities = LowesMakeupVanitiesWorkBook['Attributes']
    #Load Output Sheets

    #Load Input List
    Inputs = []
    Inputs.append(GenerateProductSheet('Bathroom Vanities','Mix',BathroomVanities,1))
    #Load Input List

    #Load Output List
    Outputs = []
    Outputs.append(GenerateProductSheet('Bathroom Vanities with Tops','Bathroom Vanities with Tops',BathroomVanitieswithTops,5))
    Outputs.append(GenerateProductSheet('Bathroom Vanities without Tops','Bathroom Vanities without Tops',BathroomVanitieswithoutTops,5))
    ###Outputs.append(GenerateProductSheet('Makeup Vanities','Makeup Vanities',MakeupVanities,5))
    #Load Output List

    #Mapping
    Inputs = Map(Inputs)
    Outputs = Map(Outputs)
    #Mapping

    AutoFill(Inputs,Outputs)
    #AutoDefault(Outputs)
    GenerateNetWeightUnpackaged(Inputs,Outputs)
    #GenerateNetWeightUnpackaged(Mapping,InputSheet,OutputSheet)

    #Save
    LowesBathroomVanitieswithTopsWorkBook.save(filename = "Bathroom Vanities with Tops.xlsx")
    LowesBathroomVanitieswithoutTopsWorkBook.save(filename = "Bathroom Vanities without Tops.xlsx")
    #LowesMakeupVanitiesWorkBook.save(filename = "Makeup Vanities.xlsx")
    #Save

if __name__ == '__main__':
    main()