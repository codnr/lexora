#Name:          Lexora EDI
#Description:   Automate EDI
#Author:        Codnr
#Version:       5
#Require:       openpyxl==2.5.14
#Notice:        Close all files used by this program before runing.

import openpyxl

def GenerateMapping():
    Mapping = []
    Data = []
    Data.append("UPC")#Name
    Data.append(["UPC"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Counter")#Name
    Data.append(["Counter"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Initiate Data Submission Date (USA)")#Name
    Data.append(["Initiate Data Submission Date (USA)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("08/26/19")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vendor Item Setup Contact Name (USA)")#Name
    Data.append(["Vendor Item Setup Contact Name (USA)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Andrey Bogan")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vendor Item Setup Contact Phone Number (USA)")#Name
    Data.append(["Vendor Item Setup Contact Phone Number (USA)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("855-453-9672")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vendor Item Setup Contact Email (USA)")#Name
    Data.append(["Vendor Item Setup Contact Email (USA)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("general@lexorahome.com")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Lowe's Merchant Email")#Name
    Data.append(["Lowe's Merchant Email"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Amanda.L.Wall@lowes.com")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("GTIN")#Name
    Data.append(["GTIN","globalTradeItemNumber (GTIN)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Highest Level GTIN")#Name
    Data.append(["Highest Level GTIN"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Product Name")#Name
    Data.append(["Title","Product Name"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Model Number")#Name
    Data.append(["Model Number","MFG Part # (OEM)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Product Type")#Name
    Data.append(["Product Type"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Information Provider GLN (USA) Initiate")#Name
    Data.append(["Information Provider GLN (USA) Initiate"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("0810014570006")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Is the Product Imported?")#Name
    Data.append(["Is the Product Imported?"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Will Lowe's Global Sourcing be the Importer of record?")#Name
    Data.append(["Will Lowe's Global Sourcing be the Importer of record?"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Supporting Sourcing Office Location")#Name
    Data.append(["Supporting Sourcing Office Location"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Shanghai")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Shipping Port Description")#Name
    Data.append(["Shipping Port Description"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("NEWARK")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Factory Name")#Name
    Data.append(["Factory Name"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Lexora")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Import Duty Rate % (USA)")#Name
    Data.append(["Import Duty Rate % (USA)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("25%")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Vendor Item Description")#Name
    Data.append(["Vendor Item Description"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Bathroom Vanities")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("MFG Brand Name")#Name
    Data.append(["MFG Brand Name"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Lexora")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Available for Direct to Consumer Delivery?")#Name
    Data.append(["Available for Direct to Consumer Delivery?"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Exclusively Available for Direct to Consumer Delivery")#Name
    Data.append(["Exclusively Available for Direct to Consumer Delivery"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Country of Origin")#Name
    Data.append(["Country of Origin"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("156")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Freight Type")#Name
    Data.append(["Freight Type"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Prepaid")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Freight Cost")#Name
    Data.append(["Freight Cost"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("0.00")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Merchandising Sub Division")#Name
    Data.append(["Merchandising Sub Division"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("155- VANITIES & MEDICINE CABINETS")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Special Order Quantity Lead Time")#Name
    Data.append(["Special Order Quantity Lead Time"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("N/A")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("List price")#Name
    Data.append(["List price","Direct to Consumer Item Cost"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Packaging Category")#Name
    Data.append(["Packaging Category"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Standard")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Estimated Annual Forecast")#Name
    Data.append(["Estimated Annual Forecast"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("2000000")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Estimated Stocked Stores")#Name
    Data.append(["Estimated Stocked Stores"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("1")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Does your product have FDC Costs? (USA)")#Name
    Data.append(["Does your product have FDC Costs? (USA)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Does your product have RDC Costs? (USA)")#Name
    Data.append(["Does your product have RDC Costs? (USA)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Does your product have Patch Store Region Costs? (USA)")#Name
    Data.append(["Does your product have Patch Store Region Costs? (USA)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("List price")#Name
    Data.append(["List price","Estimated Retail Price (USA)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Do you have MAP? (USA)")#Name
    Data.append(["Do you have MAP? (USA)"])#Alternate name list
    Data.append(None)#Location Column   
    Data.append(False)#Can Autofill
    Data.append("Yes")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Map price")#Name
    Data.append(["Map price","Minimum Advertised Price (MAP)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(True)#Can Autofills
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("hazardousMaterialClassification")#Name
    Data.append(["hazardousMaterialClassification"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("NO_MSDS_AND_NOT_REGULATED_BY_DOT_(CFR49)")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Squeeze Clamp Safe (USA)")#Name
    Data.append(["Squeeze Clamp Safe (USA)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Processing/Lead Time (Days)")#Name
    Data.append(["Processing/Lead Time (Days)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("5")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Multiple Ship From Points (USA)")#Name
    Data.append(["Multiple Ship From Points (USA)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("No")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("FOB Collect Import from Vendor (PO Net/Net) USA")#Name
    Data.append(["FOB Collect Import from Vendor (PO Net/Net) USA"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append("N/A")#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pallet Shipping Dims")#Name
    Data.append(["Pallet Shipping Dims"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Height")#Name
    Data.append(["Item Height (Inches)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Depth")#Name
    Data.append(["Item Depth (Inches)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Width")#Name
    Data.append(["Item Width (Inches)"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Pallet Weight")#Name
    Data.append(["Pallet Weight Dims"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []
    Data.append("Packaged Weight")#Name
    Data.append(["Package Weight"])#Alternate name list
    Data.append(None)#Location Column
    Data.append(False)#Can Autofill
    Data.append(None)#Default Value
    Mapping.append(Data)
    Data = []

    return Mapping

def GetMappingData(Mapping,Name):
    for Map in Mapping:
        if Map[0] == Name:
            return Map

def LoadWorkBook(FileName):
    return openpyxl.load_workbook(FileName)

def GetSpreadSheetNames(Workbook):
    return Workbook.sheetnames

def GetSpreadSheetColumnNames(Spreadsheet,Row):
    ColumnNamesList = []
    for Int in range(1,Spreadsheet.max_column+1):
        ColumnNamesList.append(Spreadsheet[GetColLetter(Int)+str(Row)].value)
    return ColumnNamesList

def GetColLetter(Col):
    return openpyxl.utils.get_column_letter(Col)

def GenerateProductSheet(Name,Type,Sheet,Row):
    return [Name,Type,Sheet,Row,GenerateMapping()]

def MapSheet(Name,Sheet,Row,Mapping):
    ColumnNamesList = GetSpreadSheetColumnNames(Sheet,Row)
    Location = 1
    for Entry in ColumnNamesList:
        for Map in Mapping:
            for Data in Map[1]:
                if Entry == Data:
                    print("Found '"+Map[0]+"' at "+str(Location)+" in '"+Name+"'")
                    Map[2] = Location
        Location += 1
    return Mapping

def Map(ProductSheet):
    for Sheet in ProductSheet:
        Sheet[4] = MapSheet(Sheet[0],Sheet[2],Sheet[3],Sheet[4])
    return ProductSheet

def CheckType(Input,Output,Location,Row):
    InputType = Input[1]
    OutputType = Output[1]
    Type = InputType
    if Type == 'Mix':
        ProductTypeColNum = GetMappingData(Input[4],'Product Type')
        CounterColNum = GetMappingData(Input[4],'Counter')
        ProductType = Input[2][GetColLetter(ProductTypeColNum[2])+str(Row+Input[3])].value
        Counter = Input[2][GetColLetter(CounterColNum[2])+str(Row+Input[3])].value
        if ProductType == 'Vanity Set':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Base Only':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity with Mirror no Counter':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity with Counter No Mirror':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Set with Side cabinet':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Vanity Set with Side cabinets':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Side Cabinet':
            Type = 'Bathroom Vanity'
        elif ProductType == 'Make-up table':
            Type = 'Makeup Vanities'
        if Type == 'Bathroom Vanity':
            if Counter == 'Yes':
                Type = 'Bathroom Vanities with Tops'
            elif Counter == 'No':
                Type = 'Bathroom Vanities without Tops'
            else:
                print('No Counter Column Input for Bathroom Vanity')
    if Type == OutputType or OutputType == 'Mix':
        return True
    else:
        return False

def WriteCopy(Input,Output,Location):
    #InputName = Input[0]
    InputSheet = Input[2]
    InputSheetOffset = Input[3]
    InputMapping = Input[4]
    #OutputName = Output[0]
    OutputSheet = Output[2]
    OutputSheetOffset = Output[3]
    OutputMapping = Output[4]
    for Row in range(1,InputSheet.max_row+1-InputSheetOffset):
        if CheckType(Input,Output,Location,Row):
            Wrote = False
            InputSheetValue = InputSheet[GetColLetter(InputMapping[Location][2])+str(Row+InputSheetOffset)].value
            for WriteRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                OutputSheetValue = OutputSheet[GetColLetter(OutputMapping[Location][2])+str(WriteRow)].value
                if OutputSheetValue == None and Wrote == False:
                    OutputSheet[GetColLetter(OutputMapping[Location][2])+str(WriteRow)].value = str(InputSheetValue)
                    Wrote = True
                    break
            if Wrote == False:
                OutputSheetValue = OutputSheet[GetColLetter(OutputMapping[Location][2])+str(OutputSheet.max_row)].value
                if OutputSheetValue != None:
                    OutputSheetValue = OutputSheet[GetColLetter(OutputMapping[Location][2])+str(OutputSheet.max_row+1)].value
                    OutputSheet[GetColLetter(OutputMapping[Location][2])+str(OutputSheet.max_row)].value = str(InputSheetValue)
            #print(InputName,InputMapping[Location][0],InputSheetValue,OutputName,OutputMapping[Location][0],OutputSheetValue)

def AutoFill(Inputs,Outputs):
    for Input in Inputs:
        for Output in Outputs:
            if Input[1] == 'Mix' or Input[1] == Output[1]:
                Location = 0
                for Map in Input[4]:
                    if Map[2] != None and Output[4][Location][2] != None and Map[3] == True:
                        print("Copy '"+Input[0]+"' "+Map[0]+" ["+GetColLetter(Map[2])+"] to '"+Output[0]+"' "+Output[4][Location][0]+" ["+GetColLetter(Output[4][Location][2])+"]")
                        WriteCopy(Input,Output,Location)
                    Location += 1

def WriteDefault(Output,Location):
    Name = Output[0]
    Sheet = Output[2]
    SheetOffset = Output[3]
    Mapping = Output[4]
    for Row in range(1,Sheet.max_row+1-SheetOffset):
        ExistingData = Sheet[GetColLetter(Mapping[Location][2])+str(Row+SheetOffset)].value
        WrittingData = Mapping[Location][4]
        if WrittingData != None:
            if ExistingData != None:
                Sheet[GetColLetter(Mapping[Location][2])+str(Row+SheetOffset)].value = WrittingData
                #print(Name+" "+Mapping[Location][0]+" Row: "+str(Row+SheetOffset)+" Overwrote Existing Data '"+str(ExistingData)+"' WrittingData '"+str(WrittingData)+"'")
            else:
                Sheet[GetColLetter(Mapping[Location][2])+str(Row+SheetOffset)].value = WrittingData
                #print(Name+" "+Mapping[Location][0]+" Row: "+str(Row+SheetOffset)+" Wrote '"+str(WrittingData)+"'")
        else:
            print(Name+" "+" Row: "+str(Row+SheetOffset)+" WrittingData is 'None'")

def AutoDefault(Outputs):
    for Output in Outputs:
        Location = 0
        for Map in Output[4]:
            if Map[2] != None and Map[4] != None:
                print("Defaulting '"+Output[0]+"' "+Map[0]+" ["+GetColLetter(Map[2])+"] to '"+str(Map[4])+"'")
                WriteDefault(Output,Location)
            Location += 1

def CalculateNetWeightpackaged(VanityWeightOnly,CounterWeightOnly,MirrorWeightOnly):
    return str(round(float(str(VanityWeightOnly).replace('N/A','0'))+float(str(CounterWeightOnly).replace('N/A','0'))+float(str(MirrorWeightOnly).replace('N/A','0')),2))

def GenerateNetWeightpackaged(Inputs,Outputs):
    for Input in Inputs:
        InputName = Input[0]
        InputSheet = Input[2]
        InputSheetOffset = Input[3]
        InputMapping = Input[4]
        for Output in Outputs:
            OutputName = Output[0]
            OutputSheet = Output[2] 
            OutputSheetOffset = Output[3]
            OutputMapping = Output[4]
            PackageWeight = GetMappingData(OutputMapping,'Packaged Weight')
            OutputUPC = GetMappingData(OutputMapping,'UPC')
            InputUPC = GetMappingData(InputMapping,'UPC')
            PalletWeight = GetMappingData(InputMapping,'Pallet Weight')
            print(PackageWeight,PalletWeight,)
            if PackageWeight != None and OutputUPC != None and InputUPC != None and PalletWeight != None:
                for OutputRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                    OutputUPCValue = OutputSheet[GetColLetter(OutputUPC[2])+str(OutputRow)].value
                    for InputRow in range(1+InputSheetOffset,InputSheet.max_row+1):
                        InputUPCValue = str(InputSheet[GetColLetter(InputUPC[2])+str(InputRow)].value)
                        if OutputUPCValue != None and InputUPCValue != None and str(OutputUPCValue) == str(InputUPCValue):
                            InputPalletWeight = InputSheet[GetColLetter(PalletWeight[2])+str(InputRow)].value
                            if InputPalletWeight != None :
                                print(str(float(InputPalletWeight) - 60).replace(".0",""))
                                CalculatedWeightpackaged = str(float(InputPalletWeight) - 60).replace(".0","")
                            else:
                                print("Error '"+OutputName+"' Row: "+str(OutputRow)+" Can't Calculate "+PackageWeight[0]+" ["+GetColLetter(PackageWeight[2])+"] "+"of '"+InputName+"' Row: "+str(InputRow))
                                CalculatedWeightpackaged = 'Missing'
                            OutputSheet[GetColLetter(PackageWeight[2])+str(OutputRow)].value = CalculatedWeightpackaged
                        else:
                            if OutputUPCValue == None:
                                print("Error '"+OutputName+"' Row: "+str(OutputRow)+" No UPC")
                                break
#end of GenerateNetWeightpackaged 

def GenerateGTIN(Outputs):
    for Output in Outputs:
        OutputName = Output[0]
        OutputSheet = Output[2]
        OutputSheetOffset = Output[3]
        OutputMapping = Output[4]
        UPC = GetMappingData(OutputMapping,'UPC')
        GTIN = GetMappingData(OutputMapping,'GTIN')
        HighestLevelGTIN = GetMappingData(OutputMapping,'Highest Level GTIN')
        if UPC != None and UPC[2] != None and GTIN and GTIN[2] != None and HighestLevelGTIN and HighestLevelGTIN[2] != None:
            print("Generating '"+OutputName+"' "+GTIN[0])
            for OutputRow in range(1+OutputSheetOffset,OutputSheet.max_row+1):
                OutputUPC = str(OutputSheet[GetColLetter(UPC[2])+str(OutputRow)].value)
                OutputGTIN = OutputSheet[GetColLetter(GTIN[2])+str(OutputRow)].value
                OutputHighestLevelGTIN = OutputSheet[GetColLetter(HighestLevelGTIN[2])+str(OutputRow)].value
                if OutputGTIN != None:
                    OutputSheet[GetColLetter(GTIN[2])+str(OutputRow)].value = '00'+OutputUPC
                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote GTIN of UPC: "+str(OutputUPC)+" to "+"Old: "+str(OutputGTIN)+" New: "+'00'+OutputUPC)
                else:
                    OutputSheet[GetColLetter(GTIN[2])+str(OutputRow)].value = '00'+OutputUPC
                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote GTIN of UPC: "+str(OutputUPC)+" to "+'00'+OutputUPC)
                if OutputHighestLevelGTIN != None:
                    OutputSheet[GetColLetter(HighestLevelGTIN[2])+str(OutputRow)].value = '00'+OutputUPC
                    #print(OutputName+" Row: "+str(OutputRow)+" Overwrote Highest Level GTIN of UPC: "+str(OutputUPC)+" to "+"Old: "+str(OutputHighestLevelGTIN)+" New: "+'00'+OutputUPC)
                else:
                    OutputSheet[GetColLetter(HighestLevelGTIN[2])+str(OutputRow)].value = '00'+OutputUPC
                    #print(OutputName+" Row: "+str(OutputRow)+" Wrote Highest Level GTIN of UPC: "+str(OutputUPC)+" to "+'00'+OutputUPC)


def main():
    #Load Input Workbooks
    LexoraSourceWorkBook = LoadWorkBook("Lexora Source.xlsx")
    #Load Input Workbooks

    #Load Output Workbooks
    #LowesBathroomVanitieswithTopsWorkBook = LoadWorkBook("Bathroom Vanities with Tops.xlsx")
    #LowesBathroomVanitieswithoutTopsWorkBook = LoadWorkBook("Bathroom Vanities without Tops.xlsx")
    InitiateWorkbook = LoadWorkBook("initiate1.xlsx")
    ###LowesMakeupVanitiesWorkBook = LoadWorkBook("Makeup Vanities.xlsx")
    #Load Output Workbooks

    #Load Input Sheets
    BathroomVanities = LexoraSourceWorkBook['Bathroom Vanities']
    #Load Input Sheets

    #Load Output Sheets
    Initiate = InitiateWorkbook['Attributes']
    #BathroomVanitieswithTops = LowesBathroomVanitieswithTopsWorkBook['Attributes']
    #BathroomVanitieswithoutTops = LowesBathroomVanitieswithoutTopsWorkBook['Attributes']
    ###MakeupVanities = LowesMakeupVanitiesWorkBook['Attributes']
    #Load Output Sheets

    #Load Input List
    Inputs = []
    Inputs.append(GenerateProductSheet('Bathroom Vanities','Mix',BathroomVanities,1))
    #Load Input List

    #Load Output List
    Outputs = []
    Outputs.append(GenerateProductSheet("initiate1",'Mix',Initiate,5))
    #Outputs.append(GenerateProductSheet('Bathroom Vanities without Tops','Bathroom Vanities without Tops',BathroomVanitieswithoutTops,5))
    #Load Output List

    #Mapping
    Inputs = Map(Inputs)
    Outputs = Map(Outputs)
    #Mapping

    #AutoFill
    AutoFill(Inputs,Outputs)
    #AutoFill

    #AutoDefault
    AutoDefault(Outputs)
    #AutoDefault

    #Generators
    GenerateNetWeightpackaged(Inputs,Outputs)
    GenerateGTIN(Outputs)
    #GenerateDimensions(Inputs,Outputs)
    #Generators

    #Save
    InitiateWorkbook.save(filename = "initiate2output.xlsx")
    #Save

if __name__ == '__main__':
    main()